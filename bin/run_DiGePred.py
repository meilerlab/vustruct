#!/usr/bin/env python
import os
import grp
import logging
from logging.handlers import RotatingFileHandler
import json
import pandas as pd
import numpy as np
from scipy import stats
import itertools
import argparse
import datetime
from scipy import stats
import seaborn
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import math
import sys
import codecs
import plotly.io as pio
import plotly.graph_objects as go
import re

ch = logging.StreamHandler()
LOGGER = logging.getLogger()
LOGGER.addHandler(ch)

log_format_string = '%(asctime)s %(levelname)-4s [%(filename)16s:%(lineno)d] %(message)s'
date_format_string = '%H:%M:%S'
log_formatter = logging.Formatter(log_format_string, date_format_string)
ch.setFormatter(log_formatter)

LOGGER.setLevel(logging.DEBUG)
ch.setLevel(logging.INFO)

# Retain Souhrid's parser setup for now
parser = argparse.ArgumentParser(description='Get digenic score and percentile for gene pair')

parser.add_argument('-g', '--genes', type=str,
                    help='genes to get digenic preds',
                    dest='genes', required=False,
                    metavar='')

parser.add_argument('--vustruct', type=str,
                    help='VUStruct input format to get digenic preds',
                    required=False,
                    metavar='')

parser.add_argument('-e', '--excel', type=str,
                    help='UDN excel file to get digenic preds',
                    dest='excel', required=False,
                    metavar='')

parser.add_argument('-c', '--csv', type=str,
                    help='gene, tscript, cDNA variant, protein variant, zyg, inh \nin csv or tsv format to get digenic preds',
                    dest='csv', required=False,
                    metavar='')

parser.add_argument('-p', '--pairs', type=str,
                    help='pairs to get digenic preds',
                    dest='pairs', required=False,
                    metavar='')

parser.add_argument('-n', '--name', type=str,
                    help='analysis name or ID',
                    dest='name', required=False,
                    metavar='')

parser.add_argument('-m', '--model', type=str,
                    help='predictive model to be usd',
                    dest='model', required=False,
                    default='unaffected no gene overlap',
                    metavar='')

parser.add_argument('-d', '--outdir', type=str,
                    help='user output dir; default to current directory',
                    dest='dir', required=False,
                    metavar='')

parser.add_argument('-u', '--uniquekey', type=str,
                    help='unique key',
                    dest='uniquekey', required=False,
                    metavar='')

## parse user input
args = parser.parse_args()


LOGGER.info("Loading genes and predictions")

# encoding=utf8

# sys.stdout = codecs.getwriter("ASCII")(sys.stdout.detach())

now = datetime.datetime.now()
month = str(now.strftime("%b"))
day = str(now.strftime("%d"))
year = str(now.strftime("%y"))

inh_colors = {'de novo': '#fdb863',
              'bi-parental': '#5aae61',
              'maternal': '#fff7bc',
              'paternal': '#9ecae1',
              'X-linked': 'black',
              'homozygous': 'blue',
              'comp. het.': '#a65628',
              'NA': '#969696',
              }

three_aa_to_one_aa = {'ALA': 'A',
                      'ARG': 'R',
                      'ASN': 'N',
                      'ASP': 'D',
                      'CYS': 'C',
                      'GLN': 'Q',
                      'GLU': 'E',
                      'GLY': 'G',
                      'HIS': 'H',
                      'ILE': 'I',
                      'LEU': 'L',
                      'LYS': 'K',
                      'MET': 'M',
                      'PHE': 'F',
                      'PRO': 'P',
                      'SER': 'S',
                      'THR': 'T',
                      'TRP': 'W',
                      'TYR': 'Y',
                      'VAL': 'V',
                      }


out_root_folder = '/nobackup/vgi/mothcw/data/mukhes1/digepred/results'
dependencies_location = '/nobackup/vgi/mothcw/data/mukhes1/digepred/data'
current_folder = os.getcwd()
out_dir = ''

def load_json_from_dependencies(subdir: str, dictionary_filename):
    _load_filename=os.path.join(dependencies_location, subdir, dictionary_filename)
    returned_dict = {}
    with open(_load_filename) as json_f:
        returned_dict = json.load(json_f)
    return returned_dict

LOGGER.info("Loading name convertors: refseq_to_gene_dict, alias_to_approved_name_dict, approved_dict_to_alias_dict");
refseq_to_gene_dict = load_json_from_dependencies( 'name_convertors', 'refseq_to_hgnc.json')
alias_to_approved_name_dict = load_json_from_dependencies( 'name_convertors', 'alias_to_approved_gene_symbol_dict.json')
approved_dict_to_alias_dict = load_json_from_dependencies( 'name_convertors', 'approved_symbol_to_alias_dict.json')

LOGGER.info("Loading networks: pair_to_int_code, interaction_details, int_code_to_pub_code, pub_id_details")
pair_to_int_code_dict = load_json_from_dependencies( 'networks', 'interaction_to_int_code_dict.json')
interaction_details_dict = load_json_from_dependencies( 'networks', 'interaction_code_details.json')
int_code_to_pub_code_dict = load_json_from_dependencies( 'networks', 'int_code_to_pub_id.json')
pub_id_details_dict = load_json_from_dependencies( 'networks', 'pub_id_to_publication_details.json')

LOGGER.info("Loading pathways: kegg_gene_to_codes, kegg_code_to_name, reactime_gene_to_codes, reactime_code_to_name")
kegg_gene_to_codes_dict = load_json_from_dependencies( 'pathways', 'kegg_gene_to_path_codes.json')
kegg_code_to_name_dict = load_json_from_dependencies( 'pathways', 'kegg_path_code_to_name.json')
reactome_gene_to_codes_dict = load_json_from_dependencies( 'pathways', 'reactome_gene_to_path_codes.json')
reactome_code_to_name_dict = load_json_from_dependencies( 'pathways', 'reactome_path_code_to_name.json')

LOGGER.info("Loading phenotypes: hpo_gene_to_codes, pheno_code_to_name, pheno_tree")
pheno_gene_to_codes_dict = load_json_from_dependencies( 'phenotypes', 'hpo_gene_to_code.json')
pheno_code_to_name_dict = load_json_from_dependencies( 'phenotypes', 'updated_HPO_code_to_name_dict.json')
pheno_tree = load_json_from_dependencies( 'phenotypes', 'HPO_network_directed_all_shortest_path_lengths.json')

LOGGER.info("Loading coexpression: id_to_broad_tissue, id_to_narrow_tissue")
gene_to_mean_TPM_by_tissue_dict = load_json_from_dependencies( 'coexpression', 'gene_to_mean_TPM_by_tissue_dict.json')
id_to_broad_tissue_dict = load_json_from_dependencies( 'coexpression', 'id_to_broad_tissue_dict.json')
id_to_narrow_tissue_dict = load_json_from_dependencies( 'coexpression', 'id_to_narrow_tissue_dict.json')

edge_dicts = dict()
edge_dicts['ppi'] = load_json_from_dependencies( 'networks', 'ppi_all_edges_dict.json')
edge_dicts['pwy'] = load_json_from_dependencies( 'networks', 'pwy_all_edges_dict.json')
edge_dicts['txt'] = load_json_from_dependencies( 'networks', 'txt_all_edges_dict.json')

csvs_path = os.path.join(dependencies_location, 'digepred_scores', 'csvs')
new_scores_csv_path = os.path.join(dependencies_location, 'digepred_scores', 'new_scores')

_held_out_test_filename = os.path.join(dependencies_location, 'digepred_scores', 'held_out_test_digenic_feat_pred.csv')
digenic_csv_df = pd.read_csv(_held_out_test_filename)
LOGGER.info("Read %d held_out_test rows from %s" % (len(digenic_csv_df), _held_out_test_filename))
_trained_digenic_filename = os.path.join(dependencies_location, 'digepred_scores', 'train_digenic_feat_pred.csv')
_trained_digenic_df = pd.read_csv(_trained_digenic_filename)
LOGGER.info("Read %d trained digenic feature rows from %s" % (len(_trained_digenic_df), _trained_digenic_filename))

digenic_csv_df = pd.concat([digenic_csv_df, _trained_digenic_df], sort=False, ignore_index=True)

## read CSVs with DiGePred scores

csv_files = []
for r, d, f in os.walk(csvs_path):
    for file in sorted(f):
        csv_files.append(os.path.join(r, file))

new_score_csv_files = []
for r, d, f in os.walk(new_scores_csv_path):
    for file in sorted(f):
        new_score_csv_files.append(os.path.join(r, file))

pred_value_percentile_dict = json.load(open(os.path.join(dependencies_location, 'digepred_scores',
                                                           'all_human_gene_pairs_pred_value_percentile_dict.json')))
percentile_pred_value_dict = json.load(open(os.path.join(dependencies_location, 'digepred_scores',
                                                           'all_human_gene_pairs_percentile_pred_value_dict.json')))

pair_to_file_num_dict = json.load(open(os.path.join(dependencies_location, 'digepred_scores',
                                                    'pair_to_file_num_dict.json')))
file_num_to_pair_dict = json.load(open(os.path.join(dependencies_location, 'digepred_scores',
                                                    'file_to_num_to_pair_dict.json')))

out_dir = ''
model = args.model
if args.dir is not None:
    usr_out_dir = args.dir

    print('User specified output directory: {}'.format(str(usr_out_dir)))

    if not os.path.isdir(usr_out_dir):
        os.mkdir(usr_out_dir)
        print('Folder created: {}'.format(usr_out_dir))
    else:
        print('Folder exists: {}'.format(usr_out_dir))

    out_dir = usr_out_dir
else:
    out_dir = current_folder
    print('output directory not specified; current directory is output directory:\n{n}'.format(n=current_folder))

# We need to set the project name
# If not set on the command line with -n/--name then see if we can fish it from other arguments
if args.name is not None:
    project_name = args.name
    print('User specified project name: {n}'.format(n=project_name))

elif args.excel is not None:
    # Fish the project name from the excel file
    excel_file_name = args.excel

    wb = load_workbook(filename=excel_file_name)
    ws = wb.active
    parsed_excel_file = []
    for row in ws.iter_rows():
        parsed_excel_file.append([cell.value for cell in row])

    n = parsed_excel_file[0][0].find('UDN')
    case_id = str(parsed_excel_file[0][0][n:n + 9])
    project_name = case_id
    print('No name specified for project; UDN id derived from UDN excel file:\n{n}'.format(n=project_name))

else:
    # Synthesize a project name from the timestamp
    date_id = '{y}-{m}-{d}-{hr}{mi}'.format(m=month, d=day, y=year, hr=hour, mi=minute)
    out_dir = os.path.join(out_root_folder, date_id)
    project_name = date_id
    LOGGER.info('No project name or ID found; Current local time used as project name:\n{n}'.format(n=project_name))

full_csv_save_path = os.path.join(out_dir, '{n}_all_gene_pairs_digenic_metrics.csv'
                                  .format(n=project_name))

html_file_name = os.path.join(out_dir, '{n}_all_gene_pairs_summary.html'
                              .format(n=project_name))

details_json_name = os.path.join(out_dir, '{n}_all_gene_pairs_summary.json'
                                 .format(n=project_name))

LOGGER.info("out_dir set to %s", out_dir)


def get_pub_code_details(pub_id):
    if pub_id in pub_id_details_dict:
        pub_details = pub_id_details_dict[pub_id]
        f_author = pub_details['authors'].split(';')[0].split(',')[0].split(' ')[0]
        year = pub_details['year']
        journal = pub_details['journal']
        paper = '{fa} et al., {y} ({j})'.format(fa=f_author, j=journal, y=year)
    else:
        paper = 'NA'

    return paper


def sort_papers(papers):
    nk = []
    for pp in papers:
        if pp != 'NA':
            n = re.split(r'(\d+)', pp)[1]
            nk.append(n)
    sorted_papers = [p for p, y in sorted(zip(papers, nk), key=lambda pair: pair[1], reverse=True)]

    return sorted_papers


def get_int_code_details(int_code):
    ct = re.search('[a-zA-Z ]+', int_code)
    source = int_code[ct.start(): ct.end()]
    if int_code in interaction_details_dict:
        int_details = interaction_details_dict[int_code]
        i_type = int_details['interaction type'].replace('_', ' ')
        i_evidence = int_details['evidence'].replace('_', ' ').replace(
            'Cause', '').replace('-Cause', '').replace('-', '').split(':')
        pub_ids = int_details['pub ID'].split('|')
    else:
        i_type = 'NA'
        i_evidence = 'NA'
        pub_ids = []

    papers = []
    for pid in pub_ids:
        papers.append(get_pub_code_details(pid))

    sorted_papers = sort_papers(papers)

    return source, i_type, i_evidence, sorted_papers


def get_int_codes(geneA, geneB):
    if '{s},{t}'.format(s=geneA, t=geneB) in pair_to_int_code_dict:
        int_codes = pair_to_int_code_dict['{s},{t}'.format(s=geneA, t=geneB)]
    elif '{s},{t}'.format(s=geneB, t=geneA) in pair_to_int_code_dict:
        int_codes = pair_to_int_code_dict['{s},{t}'.format(s=geneB, t=geneA)]
    else:
        int_codes = []

    return int_codes


def get_int_details(geneA, geneB):
    int_codes = get_int_codes(geneA, geneB)

    int_details_dict = {'types': [],
                        'sources': [],
                        'evidences': [],
                        'references': []
                        }

    for icode in int_codes:
        source, itype, evidence, sorted_papers = get_int_code_details(icode)
        if itype != '':
            int_details_dict['types'] = list(set(int_details_dict['types']).union([itype.lower()]))
        if source != '':
            int_details_dict['sources'] = list(set(int_details_dict['sources']).union([source.lower()]))
            int_details_dict['evidences'] = list(
                set(int_details_dict['evidences']).union([e.lower() for e in evidence if e != '']))
            int_details_dict['references'].extend([sp for sp in sorted_papers if sp != ''])

    int_details_dict['references'] = sort_papers(list(set(int_details_dict['references'])))

    return int_details_dict


def get_shortest_path(geneA, geneB, net_type):
    ed = edge_dicts[net_type]
    if geneA in ed:
        if geneB in ed[geneA]:
            path = [geneA, geneB]
            return path
        else:
            for n in ed[geneA]:
                if geneB in ed[n]:
                    path = [geneA, n, geneB]
                    return path
    else:
        return []


def get_path_details(path):
    path_details_dict = dict()
    for i in range(len(path) - 1):
        ga = path[i]
        gb = path[i + 1]
        path_details_dict[(ga, gb)] = get_int_details(ga, gb)

    return path_details_dict


def get_shortest_path_details(geneA, geneB):
    int_codes = get_int_codes(geneA, geneB)
    if int_codes:
        return [geneA, geneB], {(geneA, geneB): get_int_details(geneA, geneB)}
    else:
        ppi_path = get_shortest_path(geneA, geneB, net_type='ppi')
        pwy_path = get_shortest_path(geneA, geneB, net_type='pwy')
        txt_path = get_shortest_path(geneA, geneB, net_type='txt')

        if pwy_path:
            return pwy_path, get_path_details(pwy_path)
        elif txt_path:
            return txt_path, get_path_details(txt_path)
        elif ppi_path and ppi_path[1][:2] != 'UB' and ppi_path[1] != 'RPS27A':
            return ppi_path, get_path_details(ppi_path)
        else:
            return [], {}


def get_parsed_excel_file(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active
    parsed_excel_file = []
    for row in ws.iter_rows():
        parsed_excel_file.append([cell.value for cell in row])

    return parsed_excel_file


def get_udn_id_from_excel_file(parsed_excel_file):
    n = parsed_excel_file[0][0].find('UDN')
    case_id = str(parsed_excel_file[0][0][n:n + 9])
    project_name = case_id

    return project_name


def get_inh_cols(parsed_excel_file):
    mut_marker = '\u25cf'

    label_row = []
    rels = []
    cols = {'proband': 0,
            'mom': 0,
            'dad': 0,
            }
    for r_num, row in enumerate(parsed_excel_file):
        if row[0] is not None:
            if 'gene' in str(row[0]).lower():
                label_row = row
                for c_num, col in enumerate(label_row):
                    if col is not None:
                        if 'proband' in str(col).lower():
                            cols['proband'] = c_num
                        if 'mother' in str(col).lower() or 'mom' in str(col).lower():
                            cols['mom'] = c_num
                            rels.append('mom')
                        if 'father' in str(col).lower() or 'dad' in str(col).lower():
                            cols['dad'] = c_num
                            rels.append('dad')
                break

    return cols, rels, label_row


def get_inheritance_zygosity(row, cols, rels):
    inh = ''
    zyg = ''
    proband_col = cols['proband']
    dad_col = cols['dad']
    mom_col = cols['mom']
    if row[proband_col] is not None:
        proband_inh_g = row[proband_col].split()[0].strip().rstrip()
        mom_inh_g = row[mom_col].split()[0].strip().rstrip()
        dad_inh_g = row[dad_col].split()[0].strip().rstrip()
        if '\u25cf' in proband_inh_g or '\u006f' in proband_inh_g or '-' in proband_inh_g or '\u25cb' in proband_inh_g or 'need data' in \
                row[proband_col]:
            if mom_col != 0 and mom_inh_g is not None:
                if '\u25cf' in mom_inh_g:
                    inh += 'mom'
                elif str(proband_inh_g) == str(mom_inh_g):
                    inh += 'mom'

            if dad_col != 0 and dad_inh_g is not None:
                if '\u25cf' in dad_inh_g:
                    inh += 'dad'
                elif str(proband_inh_g) == str(dad_inh_g):
                    inh += 'dad'

            if 'y' in str(proband_inh_g).lower():
                zyg += 'X-linked'
            elif str(proband_inh_g).lower().count('\u25cf') == 1:
                zyg += 'heterozygous'
            elif str(proband_inh_g).lower().count('\u25cf') >= 2:
                zyg += 'homozygous'

            if len(set(rels).intersection(['mom', 'dad'])) < 2:
                inh += 'NA'
            elif 'mom' in inh and 'dad' in inh:
                inh = 'mom, dad'
            elif inh == '':
                inh = 'de novo'

    return inh, zyg


def get_gene_tscript(row):
    if 'NM_' in row[0]:
        t = str(row[0])[str(row[0]).find('NM_'):].strip().rstrip().replace('_x000D_', '').split()[
            0].strip().rstrip()
        g = row[0].split('NM_')[0].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()

    elif 'NC_' in row[0]:
        t = str(row[0])[str(row[0]).find('NC_'):].strip().rstrip().replace('_x000D_', '').split()[
            0].strip().rstrip()
        g = row[0].split('NC_')[0].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()

    elif 'NM ' in row[0]:
        r = str(row[0]).replace('NM ', 'NM_')
        t = r[r.find('NM_'):].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()
        g = r.split('NM_')[0].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()

    elif 'NC ' in row[0]:
        r = str(row[0]).replace('NC ', 'NC_')
        t = r[r.find('NC_'):].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()
        g = r.split('NC_')[0].strip().rstrip().replace('_x000D_', '').split()[0].strip().rstrip()

    else:
        t = 'NA'
        g = str(row[0]).replace('_x000D_', '').strip().rstrip()

    return g, t


def get_chr_coord(parsed_excel_file, row, r_num):
    chr = parsed_excel_file[r_num][1]
    pos = parsed_excel_file[r_num + 1][1]
    rsid = [parsed_excel_file[r_num + 1][2] if parsed_excel_file[r_num + 1][2] is not None
            else ''][0]
    ref = [parsed_excel_file[r_num][2].split('→')[0].strip().rstrip() if
           len(parsed_excel_file[r_num]) > 2 and parsed_excel_file[r_num][2] is not None and
           '→' in parsed_excel_file[r_num][2] else ''][0]
    alt = [parsed_excel_file[r_num][2].split('→')[1].strip().rstrip() if
           len(parsed_excel_file[r_num]) > 2 and parsed_excel_file[r_num][2] is not None and
           '→' in parsed_excel_file[r_num][2] else ''][0]
    consequence = [parsed_excel_file[r_num][3] if len(parsed_excel_file[r_num]) > 3 else ''][0]

    return chr, pos, rsid, ref, alt, consequence


def get_prot_var(parsed_excel_file, row, r_num):
    if parsed_excel_file[r_num + 2][2] is not None and 'p.' in parsed_excel_file[r_num + 2][2]:
        mut_long = parsed_excel_file[r_num + 2][2]
        new_string = mut_long
        for i in range(len(mut_long) - 2):
            code = mut_long[i:i + 3]
            if code.upper() in three_aa_to_one_aa:
                one_letter_code = three_aa_to_one_aa[code.upper()]
                new_string = new_string.replace(code, one_letter_code)
        pv = new_string
    else:
        pv = ''

    return pv


def get_cdna_var(parsed_excel_file, row, r_num):
    if parsed_excel_file[r_num + 1][2] is not None and 'c.' in parsed_excel_file[r_num + 1][2]:
        cv = parsed_excel_file[r_num + 1][2]
    else:
        cv = ''

    return cv


def match_gene_tscript_names(g, t):
    d_name = g
    r_name = g
    if t != '':
        derived_g_names = [refseq_to_gene_dict[t] if t in refseq_to_gene_dict else ['']][0]
        if derived_g_names != [''] and g not in derived_g_names:
            nname = derived_g_names[0] + '_({})'.format(g)
            d_name = nname
            r_name = derived_g_names[0]

    return d_name, r_name


def check_gene_alias(gname):
    # Chris notes 2024 July - this is extremely bizarre to refer to the 'g' global - but
    # that is what Souhrid had - and we keep that
    d_name = g
    r_name = g
    if gname in alias_to_approved_name_dict and gname not in approved_dict_to_alias_dict:
        rname = alias_to_approved_name_dict[gname]

    return d_name, r_name


def check_gene_in_list(rn, dn, t, genes, tscripts):
    if t != '':
        if g in genes:
            if t in tscripts:
                return genes, tscripts
            elif t not in tscripts:
                result = True
                x = 2
                while result:
                    nname = g + '_{}'.format(x)
                    if nname not in genes:
                        genes.append(nname)
                        tscripts.append(t)
                        result = False
                    else:
                        x += 1
                return genes, tscripts
        elif g not in genes:
            if t not in tscripts:
                genes.append(g)
                tscripts.append(t)
                return genes, tscripts
            elif t in tscripts:
                result = True
                x = 2
                while result:
                    nname = t + '_{}'.format(x)
                    if nname not in tscripts:
                        genes.append(g)
                        tscripts.append(nname)
                        result = False
                    else:
                        x += 1
                return genes, tscripts
    elif g not in genes:
        genes.append(g)
        return genes, tscripts
    elif g in genes:
        return genes, tscripts


def get_relevant_digepred_score_file_nums(gene_pairs):
    genes = []
    for gp in gene_pairs:
        genes.extend(list(gp))

    file_num_dicts = {'geneA': {},
                      'geneB': {},
                      'pairs': {},
                      }

    for p in sorted(gene_pairs):
        gp = tuple(sorted([p[0], p[1]]))
        cgp = tuple(sorted([gene_disp_name_to_ref_name_dict[p[0]], gene_disp_name_to_ref_name_dict[p[1]]]))

        for pp in pair_to_file_num_dict:
            ppx = pp.replace('(', '').replace(')', '').replace('\'', '').replace(' ', '').split(',')
            fp = tuple([ppx[0], ppx[1]])
            lp = tuple([ppx[2], ppx[3]])

            if fp <= cgp <= lp:

                file_num = pair_to_file_num_dict[pp]
                if file_num not in file_num_dicts['geneA']:
                    file_num_dicts['geneA'][file_num] = []
                    file_num_dicts['geneB'][file_num] = []
                    file_num_dicts['pairs'][file_num] = []
                file_num_dicts['pairs'][file_num].append(cgp)
                file_num_dicts['geneA'][file_num].append(cgp[0])
                file_num_dicts['geneB'][file_num].append(cgp[1])

    return file_num_dicts


def get_sel_genepairs_df(file_num_dicts, gene_pairs):
    sel_gp_csv_df = pd.DataFrame()

    for fnum in file_num_dicts['pairs']:
        pairs = file_num_dicts['pairs'][fnum]
        num_csv_df = pd.read_csv([x for x in csv_files if '_{}.csv'.format(fnum) in x][0])
        gp_csv_df = num_csv_df[num_csv_df['gene A'].isin(file_num_dicts['geneA'][fnum]) &
                               num_csv_df['gene B'].isin(file_num_dicts['geneB'][fnum])]

        sel_gp_csv_df = pd.concat([sel_gp_csv_df, gp_csv_df], sort=False, ignore_index=True)

    genes = []
    for gp in gene_pairs:
        genes.extend(list(gp))

    genes = list(set(genes))

    for fpath in new_score_csv_files:
        csv_df = pd.read_csv(fpath)
        ngp_csv_df = csv_df[(csv_df['gene A'].isin(genes) & csv_df['gene B'].isin(genes))]

        sel_gp_csv_df = pd.concat([sel_gp_csv_df, ngp_csv_df], sort=False, ignore_index=True)

    return sel_gp_csv_df

# These global data structures are the core of our work...
genes = []
pairs = []

gene_var_details_dict = dict()
gene_disp_name_to_ref_name_dict = dict()
gene_ref_name_to_disp_name_dict = dict()


def get_digepred_scores_df(sel_gp_csv_df):
    percentiles = []
    pwy_sims = []
    pheno_sims = []
    ppi_dists = []
    pwy_dists = []
    lit_dists = []
    coex_ranks = []
    genes_A = []
    genes_B = []
    tscripts_A = []
    tscripts_B = []
    zygs_A = []
    zygs_B = []
    cdna_vars_A = []
    cdna_vars_B = []
    prot_vars_A = []
    prot_vars_B = []
    consqs_A = []
    consqs_B = []
    inhs_A = []
    inhs_B = []
    gp_inhs = []
    preds = []

    for i in sorted(sel_gp_csv_df.index):

        ga = gene_ref_name_to_disp_name_dict[sel_gp_csv_df.iloc[i]['gene A']]
        gb = gene_ref_name_to_disp_name_dict[sel_gp_csv_df.iloc[i]['gene B']]

        genes_A.append(ga)
        genes_B.append(gb)

        pred = '%0.3g' % float(sel_gp_csv_df.iloc[i]['{m} score'.format(m=model)])
        preds.append(pred)
        pwy_sims.append(sel_gp_csv_df.iloc[i]['pathway similarity'] * 100)
        pheno_sims.append(sel_gp_csv_df.iloc[i]['phenotype similarity'] * 100)
        ppi_dists.append([1. / sel_gp_csv_df.iloc[i]['PPI distance']
                          if sel_gp_csv_df.iloc[i]['PPI distance'] > 0
                          else 0][0])
        pwy_dists.append([1. / sel_gp_csv_df.iloc[i]['pathway distance']
                          if sel_gp_csv_df.iloc[i]['pathway distance'] > 0
                          else 0][0])
        lit_dists.append([1. / sel_gp_csv_df.iloc[i]['literature distance']
                          if sel_gp_csv_df.iloc[i]['literature distance'] > 0
                          else 0][0])
        coex_ranks.append([1. / sel_gp_csv_df.iloc[i]['mutual co-expression rank']
                           if sel_gp_csv_df.iloc[i]['mutual co-expression rank'] > 0
                           else 0][0])

        res = True
        for i in range(len(pred_value_percentile_dict) - 1):
            lower = float(list(pred_value_percentile_dict)[i])
            upper = float(list(pred_value_percentile_dict)[i + 1])
            if lower <= float(pred) < upper:
                ptile = pred_value_percentile_dict[list(pred_value_percentile_dict)[i]]
                res = False
                break

        if res:
            ptile = 99

        percentiles.append(ptile)

        if ga in gene_var_details_dict:
            tscripts_A.append(gene_var_details_dict[ga]['tscript'])
            cdna_vars_A.append(gene_var_details_dict[ga]['cdna var'])
            prot_vars_A.append(gene_var_details_dict[ga]['prot var'])
            consqs_A.append(gene_var_details_dict[ga]['consequence'])
            zygs_A.append(gene_var_details_dict[ga]['zyg'])

            ii_A = gene_var_details_dict[ga]['inh']
            zyg_A = gene_var_details_dict[ga]['zyg']

            mom_inh_A = False
            dad_inh_A = False
            de_novo_A = False

            if 'mom' in str(ii_A) or 'mother' in str(ii_A) or 'maternal' in str(ii_A):
                mom_inh_A = True
            if 'dad' in str(ii_A) or 'father' in str(ii_A) or 'paternal' in str(ii_A):
                dad_inh_A = True
            if 'de novo' in str(ii_A):
                de_novo_A = True

            inh_A = 'NA'
            if mom_inh_A or dad_inh_A:
                if mom_inh_A and dad_inh_A:
                    inh_A = 'comp het'
                elif mom_inh_A:
                    inh_A = 'maternal'
                elif dad_inh_A:
                    inh_A = 'paternal'
            elif 'NA' not in str(ii_A):
                inh_A = 'de novo'
            else:
                inh_A = 'NA'
            if de_novo_A and 'de novo' not in inh_A:
                inh_A = inh_A + ' de novo'

        else:
            ii_A = 'NA'
            inh_A = 'NA'
            zyg_A = 'NA'
            tscripts_A.append('NA')
            cdna_vars_A.append('NA')
            prot_vars_A.append('NA')
            consqs_A.append('NA')
            zygs_A.append('NA')

        if gb in gene_var_details_dict:
            tscripts_B.append(gene_var_details_dict[gb]['tscript'])
            cdna_vars_B.append(gene_var_details_dict[gb]['cdna var'])
            prot_vars_B.append(gene_var_details_dict[gb]['prot var'])
            consqs_B.append(gene_var_details_dict[gb]['consequence'])
            zygs_B.append(gene_var_details_dict[gb]['zyg'])

            ii_B = gene_var_details_dict[gb]['inh']
            zyg_B = gene_var_details_dict[gb]['zyg']

            mom_inh_B = False
            dad_inh_B = False
            de_novo_B = False

            if 'mom' in str(ii_B) or 'mother' in str(ii_B) or 'maternal' in str(ii_B):
                mom_inh_B = True
            if 'dad' in str(ii_B) or 'father' in str(ii_B) or 'paternal' in str(ii_B):
                dad_inh_B = True
            if 'de novo' in str(ii_B):
                de_novo_B = True

            inh_B = 'NA'
            if mom_inh_B or dad_inh_B:
                if mom_inh_B and dad_inh_B:
                    inh_B = 'comp het'
                elif mom_inh_B:
                    inh_B = 'maternal'
                elif dad_inh_B:
                    inh_B = 'paternal'
            elif 'NA' not in str(ii_B):
                inh_B = 'de novo'
            else:
                inh_B = 'NA'

            if de_novo_B and 'de novo' not in inh_B:
                inh_B = inh_B + ' de novo'

        else:
            ii_B = 'NA'
            inh_B = 'NA'
            zyg_B = 'NA'
            tscripts_B.append('NA')
            cdna_vars_B.append('NA')
            prot_vars_B.append('NA')
            consqs_B.append('NA')
            zygs_B.append('NA')

        gp_inh = 'NA'

        inhzygA = inh_A + zyg_A
        inhzygB = inh_B + zyg_B

        iza = ' '
        if 'de novo' in inhzygA:
            iza += '<i>de novo</i>,'
        if 'homozygous' in inhzygA:
            iza += 'homozygous,'
        if 'comp het' in inhzygA:
            iza += 'compound het.,'
        if 'X-linked' in inhzygA:
            iza += 'X-linked,'
        if 'maternal' in inhzygA and 'paternal' not in inhzygA:
            iza += 'maternal,'
        if 'paternal' in inhzygA and 'maternal' not in inhzygA:
            iza += 'paternal,'

        izb = ' '
        if 'de novo' in inhzygB:
            izb += '<i>de novo</i>,'
        if 'homozygous' in inhzygB:
            izb += 'homozygous,'
        if 'comp het' in inhzygB:
            izb += 'compound het.,'
        if 'X-linked' in inhzygB:
            izb += 'X-linked,'
        if 'maternal' in inhzygB and 'paternal' not in inhzygB:
            izb += 'maternal,'
        if 'paternal' in inhzygB and 'maternal' not in inhzygB:
            izb += 'paternal,'

        if 'de novo' in iza or 'de novo' in izb:
            iz = 'de novo'
            u = 'unique in proband'
        elif 'homozygous' in iza or 'homozygous' in izb:
            iz = 'homozygous'
            u = 'unique in proband'
        elif 'compound het.' in iza or 'compound het.' in izb:
            iz = 'compound het.'
            u = 'unique in proband'
        elif 'maternal' in iza and 'paternal' in izb:
            iz = 'bi-parental'
            u = 'unique in proband'
        elif 'paternal' in iza and 'maternal' in izb:
            iz = 'bi-parental'
            u = 'unique in proband'
        elif 'X-linked' in iza or 'X-linked' in izb:
            iz = 'X-linked'
            u = 'unique in proband'
        elif 'maternal' in iza and 'maternal' in izb:
            iz = 'maternal'
            u = ''
        elif 'paternal' in iza and 'paternal' in izb:
            iz = 'paternal'
            u = ''
        else:
            iz = 'NA'
            u = ''

        inhs_A.append(iza[:-1])
        inhs_B.append(izb[:-1])
        if u != '':
            gp_inhs.append('{izz} ({uu})'.format(izz=iz, uu=u))
        else:
            gp_inhs.append(iz)

    dig_df = pd.DataFrame(index=range(1, len(sel_gp_csv_df.index) + 1))
    dig_df['gene A'] = genes_A
    dig_df['gene B'] = genes_B
    dig_df['digenic score'] = preds
    dig_df['pair inheritance'] = gp_inhs
    dig_df['percentile'] = percentiles
    dig_df['transcript A'] = tscripts_A
    dig_df['cDNA change A'] = cdna_vars_A
    dig_df['protein change A'] = prot_vars_A
    dig_df['consq. A'] = consqs_A
    dig_df['zygosity A'] = zygs_A
    dig_df['inheritance A'] = inhs_A
    dig_df['transcript B'] = tscripts_B
    dig_df['cDNA change B'] = cdna_vars_B
    dig_df['protein change B'] = prot_vars_B
    dig_df['consq. B'] = consqs_B
    dig_df['zygosity B'] = zygs_B
    dig_df['inheritance B'] = inhs_B
    dig_df['pathway similarity'] = pwy_sims
    dig_df['phenotype similarity'] = pheno_sims
    dig_df['PPI distance'] = ppi_dists
    dig_df['pathway distance'] = pwy_dists
    dig_df['literature distance'] = lit_dists
    dig_df['mutual co-expression rank'] = coex_ranks

    return dig_df


def get_hover_text(df):
    vdf = pd.DataFrame(index=df.index,
                       columns=df.columns,
                       data=[[str(df[c][r])
                              for c in df.columns]
                             for r in df.index])

    vdf = vdf.replace('nan', 'NA')

    gp_inh_dict, inh_gp_dict = get_inh_dict(dig_info_dict)

    for row in df.index:
        for col in df.columns:
            if row != col and vdf[col][row] != 'NA':
                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                        gb=tuple(sorted([row, col]))[1])
                ds = dig_info_dict[ip]['DiGePred']['score']
                dp = dig_info_dict[ip]['DiGePred']['percentile']
                path = dig_info_dict[ip]['path']
                dist = dig_info_dict[ip]['distance']

                iz = dig_info_dict[ip]['inheritance']
                iza = dig_info_dict[ip]['variants'][row]['inh']
                izb = dig_info_dict[ip]['variants'][col]['inh']
                consqa = dig_info_dict[ip]['variants'][row]['consq']
                if 'missense' in consqa:
                    prota = dig_info_dict[ip]['variants'][row]['prot']
                    if prota != "":
                        consqa = prota
                consqb = dig_info_dict[ip]['variants'][col]['consq']
                if 'missense' in consqb:
                    protb = dig_info_dict[ip]['variants'][col]['prot']
                    if protb != "":
                        consqb = protb

                ds = '% 0.2g' % float(dig_info_dict[ip]['DiGePred']['score'])
                disp_text = '<br>DiGePred score: {ds}<br>{dp} percentile'.format(ds=ds, dp=dp)
                disp_text += '<br>{g} ({vt} | {iz})'.format(g=row, vt=consqa, iz=iza)
                disp_text += '<br>{g} ({vt} | {iz})'.format(g=col, vt=consqb, iz=izb)
                if path:
                    if len(path) == 2:
                        disp_text += '<br>direct interaction'
                        itypes = []
                        for x in dig_info_dict[ip]['interactions'][ip]['types']:
                            itypes.extend(x.split(' '))
                        itypes = set(itypes)
                        isources = dig_info_dict[ip]['interactions'][ip]['sources']
                        irefs = dig_info_dict[ip]['interactions'][ip]['references']
                        disp_text += '<br>    ' + ';'.join(isources)
                        if itypes:
                            disp_text += '<br>    ' + ';'.join(itypes)
                        disp_text += '<br>    ' + 'References ({n}):'.format(n=len(irefs))
                        disp_text += '<br>    ' + '<br>        '.join(irefs[:np.minimum(len(irefs), 4)])
                    elif len(path) > 2:
                        disp_text += '<br>indirect interaction (' + '-'.join(path) + ')'
                        for gpx in dig_info_dict[ip]['interactions']:
                            gx = gpx.split(',')
                            disp_text += '<br>    ' + gx[0] + '-' + gx[1]
                            ptypes = []
                            for x in dig_info_dict[ip]['interactions'][gpx]['types']:
                                ptypes.extend(x.split(' '))
                            ptypes = set(ptypes)
                            psources = dig_info_dict[ip]['interactions'][gpx]['sources']
                            prefs = dig_info_dict[ip]['interactions'][gpx]['references']
                            disp_text += '<br>    ' + ';'.join(psources)
                            if ptypes:
                                disp_text += '<br>    ' + ';'.join(ptypes)
                            disp_text += '<br>    ' + 'References ({n}):'.format(n=len(prefs))
                            disp_text += '<br>    ' + '<br>        '.join(prefs[:np.minimum(len(prefs), 2)])

                pwy_sim = dig_info_dict[ip]['pathways']['similarity']
                pheno_sim = dig_info_dict[ip]['phenotypes']['similarity']
                coex = dig_info_dict[ip]['coexpression']['rank']

                if pwy_sim > 0:
                    j_pwy = dig_info_dict[ip]['pathways']['similarity']
                    pCommon = dig_info_dict[ip]['pathways']['common']['names']
                    nCommon = dig_info_dict[ip]['pathways']['# common']
                    nTotal = dig_info_dict[ip]['pathways']['# total']
                    disp_text += '<br>pathway similarity: {pw}% ({nc}/{nt})'.format(nc=nCommon, nt=nTotal, pw=pwy_sim)
                    disp_text += '<br>    ' + '<br>    '.join(pCommon[:np.minimum(len(pCommon), 5)])
                if pheno_sim > 0:
                    j_pheno = dig_info_dict[ip]['phenotypes']['similarity']
                    com_phenos = dig_info_dict[ip]['phenotypes']['common']['codes']
                    nCommon = dig_info_dict[ip]['phenotypes']['# common']
                    nTotal = dig_info_dict[ip]['phenotypes']['# total']

                    if nTotal > 0:
                        disp_text += '<br>phenotype similarity: {pw}% ({nc}/{nt})'.format(nc=nCommon, nt=nTotal,
                                                                                          pw=pheno_sim)

                        if len(com_phenos) > 5:
                            npx = []
                            for p in com_phenos:
                                if p in pheno_tree:
                                    npx.append(list(pheno_tree[p])[
                                                   [-4 if len(pheno_tree[p]) > 4 else 0][0]])

                            com_phenos = npx

                        com_pheno_names = sorted(set([pheno_code_to_name_dict[c]
                                                      for c in com_phenos if
                                                      c in pheno_code_to_name_dict and
                                                      'obsolete' not in pheno_code_to_name_dict[
                                                          c].lower()]))

                        disp_text += '<br>    ' + '<br>    '.join(com_pheno_names[:np.minimum(len(com_pheno_names), 5)])
                if coex != 'NA':
                    coex = int(coex)
                    coex_tissues = dig_info_dict[ip]['coexpression']['tissues']
                    disp_text += '<br>mutual co-expression rank: {pw}'.format(pw=coex)
                    disp_text += '<br>    ' + '<br>    '.join(coex_tissues[:np.minimum(len(coex_tissues), 5)])

                vdf[row][col] = disp_text
                vdf[col][row] = disp_text

    return vdf


def get_digenic_metrics_csv(gene_pairs):
    gene_pairs = set([tuple(sorted(gp)) for gp in gene_pairs])

    file_num_dicts = get_relevant_digepred_score_file_nums(gene_pairs)
    sel_gp_csv_df = get_sel_genepairs_df(file_num_dicts, gene_pairs)

    dig_df = get_digepred_scores_df(sel_gp_csv_df)

    print('\nDigenic metrics computed for all gene pairs\n')
    print(dig_df)

    dig_df.to_csv(full_csv_save_path, sep=',', header=True, index=False)

    print('output csv filename: {}'.format(full_csv_save_path))

    return dig_df, sel_gp_csv_df


def plot_html(df, dig_info_dict):
    genes = sorted(list(set(df['gene A']).union(df['gene B'])))

    dfs = {'DiGePred': {}, 'systems biology': {}}

    dfs['DiGePred']['score'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes), dtype=float)
    dfs['DiGePred']['pcile'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes), dtype=float)
    dfs['systems biology']['dir. int.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                       dtype=float)
    dfs['systems biology']['indir. int.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                         dtype=float)
    dfs['systems biology']['pwy sim.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                      dtype=float)
    dfs['systems biology']['pheno sim.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                        dtype=float)
    dfs['systems biology']['coex.'] = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes),
                                                   dtype=float)

    data_ranges = {'DiGePred': {}, 'systems biology': {}}
    color_scales = {'DiGePred': {}, 'systems biology': {}}

    data_ranges['DiGePred']['score'] = [0.01, 0.1, 0.5]
    data_ranges['DiGePred']['pcile'] = []
    data_ranges['systems biology']['dir. int.'] = [1]
    data_ranges['systems biology']['indir. int.'] = [1]
    data_ranges['systems biology']['pwy sim.'] = [0.01, 10, 50]
    data_ranges['systems biology']['pheno sim.'] = [0.01, 10, 50]
    data_ranges['systems biology']['coex.'] = [2000, 200, 20]

    trace_names = {'DiGePred': {}, 'systems biology': {}}
    trace_names['DiGePred']['score'] = 'DiGePred score'
    trace_names['DiGePred']['pcile'] = 'percentile of DiGePred score'
    trace_names['systems biology']['dir. int.'] = 'direct'
    trace_names['systems biology']['indir. int.'] = 'indirect'
    trace_names['systems biology']['pwy sim.'] = 'KEGG and Reactome pathway similarity'
    trace_names['systems biology']['pheno sim.'] = 'HPO phenotype similarity'
    trace_names['systems biology']['coex.'] = 'mutual co-expression rank'

    inh_df = pd.DataFrame(index=sorted(genes, reverse=True), columns=sorted(genes), dtype=str)

    for i in df.index:
        ga = df['gene A'][i]
        gb = df['gene B'][i]
        score = df['digenic score'][i]

        dfs['DiGePred']['score'].at[ga, gb] = score
        dfs['DiGePred']['score'].at[gb, ga] = score

        dfs['DiGePred']['pcile'].at[ga, gb] = df['percentile'][i]
        dfs['DiGePred']['pcile'].at[gb, ga] = df['percentile'][i]

        dfs['systems biology']['dir. int.'].at[ga, gb] = [1 if df['PPI distance'][i] == 1 or
                                                               df['pathway distance'][i] == 1 or
                                                               df['literature distance'][i] == 1 else np.nan][0]
        dfs['systems biology']['dir. int.'].at[gb, ga] = [1 if df['PPI distance'][i] == 1 or
                                                               df['pathway distance'][i] == 1 or
                                                               df['literature distance'][i] == 1 else np.nan][0]

        dfs['systems biology']['indir. int.'].at[ga, gb] = [1 if df['PPI distance'][i] == 2 or
                                                                 df['pathway distance'][i] == 2 or
                                                                 df['literature distance'][i] == 2 else np.nan][0]
        dfs['systems biology']['indir. int.'].at[gb, ga] = [1 if df['PPI distance'][i] == 2 or
                                                                 df['pathway distance'][i] == 2 or
                                                                 df['literature distance'][i] == 2 else np.nan][0]

        dfs['systems biology']['pwy sim.'].at[ga, gb] = df['pathway similarity'][i]
        dfs['systems biology']['pwy sim.'].at[gb, ga] = df['pathway similarity'][i]

        dfs['systems biology']['pheno sim.'].at[ga, gb] = df['phenotype similarity'][i]
        dfs['systems biology']['pheno sim.'].at[gb, ga] = df['phenotype similarity'][i]

        dfs['systems biology']['coex.'].at[ga, gb] = [df['mutual co-expression rank'][i] if
                                                      df['mutual co-expression rank'][i] > 0 else np.nan][0]
        dfs['systems biology']['coex.'].at[gb, ga] = [df['mutual co-expression rank'][i] if
                                                      df['mutual co-expression rank'][i] > 0 else np.nan][0]
        inh_df.at[ga, gb] = df['pair inheritance'][i]
        inh_df.at[gb, ga] = df['pair inheritance'][i]

    for g in genes:
        dfs['DiGePred']['score'][g][g] = np.nan
        dfs['DiGePred']['pcile'][g][g] = np.nan
        dfs['systems biology']['dir. int.'][g][g] = np.nan
        dfs['systems biology']['indir. int.'][g][g] = np.nan
        dfs['systems biology']['pwy sim.'][g][g] = np.nan
        dfs['systems biology']['pheno sim.'][g][g] = np.nan
        dfs['systems biology']['coex.'][g][g] = np.nan

    best_vals = {'DiGePred': {}, 'systems biology': {}}
    best_vals['DiGePred']['score'] = 0.5
    best_vals['systems biology']['dir. int.'] = 1
    best_vals['systems biology']['indir. int.'] = 1
    best_vals['systems biology']['pwy sim.'] = 10
    best_vals['systems biology']['pheno sim.'] = 10
    best_vals['systems biology']['coex.'] = 500

    color_scales = {'DiGePred': {}, 'systems biology': {}}

    fig = go.Figure()
    datas = {}
    set_details = []
    values = []
    cutoffs = []
    i_fig = 0
    for k in dfs:
        datas[k] = {}
        for s in dfs[k]:
            datas[k][s] = {}

    for k in dfs:
        for s in dfs[k]:
            for i_s, step in enumerate(data_ranges[k][s]):

                color_scales['DiGePred']['score'] = [[0.0, 'white'], [1, '#3f007d']]
                color_scales['DiGePred']['pcile'] = [[0.0, 'white'], [1, '#3f007d']]
                color_scales['systems biology']['dir. int.'] = [[0.0, 'white'], [1.0, '#3f007d']]
                color_scales['systems biology']['indir. int.'] = [[0.0, 'white'], [1.0, '#3f007d']]
                color_scales['systems biology']['pwy sim.'] = [[0.0, 'white'], [1, '#3f007d']]
                color_scales['systems biology']['pheno sim.'] = [[0.0, 'white'], [1, '#3f007d']]
                color_scales['systems biology']['coex.'] = [[0.0, '#3f007d'], [1.0, 'white']]

                wide_df_i = pd.DataFrame(index=[i for i in dfs[k][s].index],
                                         data=dfs[k][s].values,
                                         columns=[c for c in dfs[k][s].columns])
                z = wide_df_i.values
                plot_df = wide_df_i.copy()

                vdf = pd.DataFrame(index=plot_df.index,
                                   columns=plot_df.columns,
                                   data=[[str(wide_df_i[c][r])
                                          for c in plot_df.columns]
                                         for r in plot_df.index])

                vdf = vdf.replace('nan', 'NA')

                if s == 'dir. int.':
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            v = plot_df[col][row]
                            if row != col and v == 1:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])

                                disp_text = 'direct interaction'
                                disp_text += '<br>' + ';'.join(dig_info_dict[ip]['interactions'][ip]['types'])
                                disp_text += '<br>' + '<br>'.join(dig_info_dict[ip]['interactions'][ip]['evidences'])
                                disp_text += '<br>' + ';'.join(dig_info_dict[ip]['interactions'][ip]['sources'])
                                if len(dig_info_dict[ip]['interactions'][ip]['references']) > 7:
                                    disp_text += '<br>' + '<br>'.join(
                                        dig_info_dict[ip]['interactions'][ip]['references'][:7])
                                else:
                                    disp_text += '<br>' + '<br>'.join(
                                        dig_info_dict[ip]['interactions'][ip]['references'])

                                vdf[row][col] = disp_text
                                vdf[col][row] = disp_text

                                ds = dig_info_dict[ip]['DiGePred']['score']
                                ds_text = '<br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text

                elif s == 'indir. int.':
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            v = plot_df[col][row]
                            if row != col and v == 1:
                                disp_text = ''
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])
                                path = dig_info_dict[ip]['path']
                                if path and len(path) > 2:
                                    disp_text = 'indirect interaction' + '<br>' + '-'.join(path)
                                    for gpx in dig_info_dict[ip]['interactions']:
                                        gx = gpx.split(',')
                                        disp_text += '<br>' + gx[0] + '-' + gx[1]
                                        disp_text += '<br>' + ';'.join(dig_info_dict[ip]['interactions'][gpx]['types'])
                                        disp_text += '<br>' + ';'.join(
                                            dig_info_dict[ip]['interactions'][gpx]['sources'])
                                        if len(dig_info_dict[ip]['interactions'][gpx]['references']) > 3:
                                            disp_text += '<br>' + '<br>'.join(
                                                dig_info_dict[ip]['interactions'][gpx]['references'][:3])
                                        else:
                                            disp_text += '<br>' + '<br>'.join(
                                                dig_info_dict[ip]['interactions'][gpx]['references'])
                                else:
                                    plot_df[col][row] = np.nan
                                    plot_df[row][col] = np.nan

                                vdf[row][col] = disp_text
                                vdf[col][row] = disp_text

                                ds = dig_info_dict[ip]['DiGePred']['score']
                                ds_text = '<br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text
                            else:
                                vdf[col][row] = 'NA'
                                vdf[row][col] = 'NA'

                if 'pwy' in s:
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            disp_text = ''
                            if row != col:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])
                                if ip in dig_info_dict:
                                    j_pwy = dig_info_dict[ip]['pathways']['similarity']
                                    pCommon = dig_info_dict[ip]['pathways']['common']['names']
                                    nCommon = dig_info_dict[ip]['pathways']['# common']
                                    nTotal = dig_info_dict[ip]['pathways']['# total']
                                    ds = dig_info_dict[ip]['DiGePred']['score']
                                else:
                                    j_pwy = 0
                                    pCommon = []
                                    nCommon = 0
                                    nTotal = 0
                                    ds = 0

                                if nTotal > 0:

                                    disp_text += '{v} %'.format(v=j_pwy)
                                    disp_text += '<br>' + '({nc}/{nt})'.format(nc=nCommon,
                                                                               nt=nTotal)
                                    disp_text += '<br>' + '<br>'.join(pCommon)
                                else:
                                    disp_text += 'NA'

                                vdf[row][col] = disp_text
                                vdf[col][row] = disp_text

                                ds_text = '<br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text
                            else:
                                vdf[col][row] = 'NA'
                                vdf[row][col] = 'NA'
                if 'pheno' in s:
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            disp_text = ''
                            if row != col:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])
                                if ip in dig_info_dict:
                                    j_pheno = dig_info_dict[ip]['phenotypes']['similarity']

                                    com_phenos = dig_info_dict[ip]['phenotypes']['common']['codes']
                                    nCommon = dig_info_dict[ip]['phenotypes']['# common']
                                    nTotal = dig_info_dict[ip]['phenotypes']['# total']
                                    ds = dig_info_dict[ip]['DiGePred']['score']
                                else:
                                    j_pheno = 0
                                    com_phenos = []
                                    nCommon = 0
                                    nTotal = 0
                                    ds = 0

                                if nTotal > 0:
                                    disp_text += '{v} %'.format(v=j_pheno)
                                    disp_text += '<br>' + '({nc}/{nt})'.format(nc=nCommon,
                                                                               nt=nTotal)

                                    if len(com_phenos) > 10:
                                        npx = []
                                        for p in com_phenos:
                                            if p in pheno_tree:
                                                npx.append(list(pheno_tree[p])[
                                                               [-4 if len(pheno_tree[p]) > 4 else 0][0]])

                                        com_phenos = npx

                                    com_pheno_names = sorted(set([pheno_code_to_name_dict[c]
                                                                  for c in com_phenos if
                                                                  c in pheno_code_to_name_dict and
                                                                  'obsolete' not in pheno_code_to_name_dict[
                                                                      c].lower()]))

                                    disp_text += '<br>' + '<br>'.join(com_pheno_names)
                                else:
                                    disp_text += 'NA'
                                vdf[row][col] = disp_text
                                vdf[col][row] = disp_text

                                ds_text = '<br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text
                            else:
                                vdf[col][row] = 'NA'
                                vdf[row][col] = 'NA'

                if k == 'DiGePred':
                    vdf = get_hover_text(plot_df)

                if s == 'coex.':
                    for row in plot_df.index:
                        for col in plot_df.columns:
                            if row != col:
                                ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                                        gb=tuple(sorted([row, col]))[1])
                                if ip in dig_info_dict:
                                    coex_rank = dig_info_dict[ip]['coexpression']['rank']
                                    coex_tissues = dig_info_dict[ip]['coexpression']['tissues']
                                    ds = dig_info_dict[ip]['DiGePred']['score']
                                else:
                                    coex_rank = 20000
                                    coex_tissues = []
                                    ds = 0

                                vdf[row][col] = '{v}'.format(v=coex_rank)
                                vdf[col][row] = '{v}'.format(v=coex_rank)

                                disp_text = ''
                                if coex_rank != 'NA' and coex_tissues:
                                    if len(coex_tissues) > 10:
                                        disp_text += '<br>' + '<br>'.join(coex_tissues[:10])
                                    else:
                                        disp_text += '<br>' + '<br>'.join(coex_tissues)

                                vdf[col][row] += disp_text
                                vdf[row][col] += disp_text

                                ds_text = '<br>DiGePred score: {ds}'.format(ds=ds)
                                vdf[col][row] += ds_text
                                vdf[row][col] += ds_text
                            else:
                                vdf[col][row] = 'NA'
                                vdf[row][col] = 'NA'

                if s != 'coex.':
                    plot_df[plot_df.astype('float') < step] = 0
                    plot_df = plot_df.fillna(value=0)
                else:
                    plot_df[plot_df.astype('float') > step] = 2000
                    plot_df = plot_df.fillna(value=2000)
                z2 = plot_df.values
                if len(z2) == 0 or plot_df.min == 0 or plot_df.max == 2000:
                    fig.add_trace(
                        go.Heatmap(visible=False, name=trace_names[k][s], opacity=0.9, zmin=0, zmax=0,
                                   x=[c for c in plot_df.columns],
                                   y=[i for i in plot_df.index],
                                   z=z2, text=vdf,
                                   type='heatmap', hoverinfo='text+x+y', hovertemplate="<b>%{text}</b><br>" +
                                                                                       "gene A: %{x}<br>" +
                                                                                       "gene B: %{y}<br>"
                                                                                       "<extra></extra>"
                                   ,
                                   colorscale=[[0.0, 'white'], [1.0, 'white']], xperiodalignment='start',
                                   colorbar=dict(tickmode='array', tickvals=data_ranges[k][s]),
                                   )
                    )
                    datas[k][s][step] = i_fig
                    set_details.append((k, s, step))
                    i_fig += 1
                else:
                    datas[k][s][step] = z2
                    cutoffs.append(step)
                    values.append(plot_df)
                    fig.add_trace(
                        go.Heatmap(visible=False, name=trace_names[k][s], opacity=0.9,
                                   zmin=data_ranges[k][s][0], zmax=data_ranges[k][s][-1],
                                   x=[c for c in plot_df.columns],
                                   y=[i for i in plot_df.index],
                                   z=z2, text=vdf,
                                   type='heatmap', hoverinfo='text+x+y', hovertemplate="<b>%{text}</b><br>" +
                                                                                       "gene A: %{x}<br>" +
                                                                                       "gene B: %{y}<br>"
                                                                                       "<extra></extra>"
                                   ,
                                   colorscale=color_scales[k][s], xperiodalignment='start',
                                   colorbar=dict(tickmode='array', tickvals=data_ranges[k][s]),
                                   )
                    )
                    datas[k][s][step] = i_fig
                    set_details.append((k, s, step))
                    i_fig += 1

    fig.data[1].visible = True
    fig.update_traces(showscale=False)

    vis_data = {}
    vis_dicts = {}
    all_vis_dicts = []
    for i, k in enumerate(dfs):
        vis_data[k] = {}
        vis_dicts[k] = {}
        for j, s in enumerate(dfs[k]):
            vis_data[k][s] = {}
            vis_dicts[k][s] = []
            nn = 0
            for st in datas[k][s]:
                vis_data[k][s][st] = [False] * len(fig.data)
                n = datas[k][s][st]
                vis_data[k][s][st][n] = True
                if 'pwy' in s or 'pheno' in s:
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {'title': '{n} | cutoff >= {c}'.format(n=trace_names[k][s],
                                                                                           c=[st if st > 0.01 else 0][
                                                                                               0]),
                                                     'font_size': 16,
                                                     'font_family': "Arial"}],
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}],
                                              label=['\u2265 {v} %'.format(v=st) if st > 0.01 else 'ALL'][0],
                                              method='update',
                                              ))
                elif s == 'dir. int.':
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {
                                                        'title': 'direct interactions (PPI, pathway & literature-mined interactions)',
                                                        'font_size': 16,
                                                        'font_family': "Arial"}
                                                    ],
                                              label='direct',
                                              method='update',
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}]
                                              ))
                elif s == 'indir. int.':
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {
                                                        'title': 'indirect interactions (PPI, pathway & literature-mined interactions)',
                                                        'font_size': 16,
                                                        'font_family': "Arial"}
                                                    ],
                                              label='indirect',
                                              method='update',
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}]
                                              ))
                elif s != 'coex.':
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {'title': '{n} | cutoff >= {c}'.format(n=trace_names[k][s],
                                                                                           c=[st if st > 0.01 else 0][
                                                                                               0]),
                                                     'font_size': 16,
                                                     'font_family': "Arial"}],
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}],
                                              label=['\u2265 {v}'.format(v=st) if st > 0.01 else 'ALL'][0],
                                              method='update',
                                              ))
                else:
                    all_vis_dicts.append(dict(args=[{'visible': vis_data[k][s][st]},
                                                    {'title': '{n} | cutoff <= {c}'.format(n=trace_names[k][s], c=st),
                                                     'font_size': 16,
                                                     'font_family': "Arial"},
                                                    ],
                                              args2=[{'visible': [False] * len(fig.data)},
                                                     {'title': 'Please select a metric',
                                                      'font_size': 16,
                                                      'font_family': "Arial"}],
                                              label='\u2264 {v}'.format(v=st),
                                              method='update',
                                              ))

    # fig = draw_inh_col_boxes_gene_level(fig, inh_df)
    fig = draw_inh_col_boxes_pair_level(fig, inh_df)

    fig.update_layout(
        updatemenus=[
            dict(showactive=True, bgcolor='#bdbdbd', bordercolor='#636363',
                 type="buttons", font_size=10, font_family="Arial",
                 direction="right",
                 active=1,
                 pad={"t": 10}, x=0.89, y=-0.17,
                 buttons=list(all_vis_dicts),
                 ),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['de novo'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.9,
                 buttons=[dict(args=[{}], label='<i>de novo</i> variants', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['homozygous'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.8,
                 buttons=[dict(args=[{}], label='homozygous variants', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['comp. het.'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.7,
                 buttons=[dict(args=[{}], label='Compound het.', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['bi-parental'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.6,
                 buttons=[dict(args=[{}], label='variants inherited<br>from both parents', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['X-linked'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.5,
                 buttons=[dict(args=[{}], label='X-linked', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['maternal'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.4,
                 buttons=[dict(args=[{}], label='maternally inherited<br>variants only', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['paternal'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.3,
                 buttons=[dict(args=[{}], label='paternally inherited<br>variants only', method='skip')]),
            dict(showactive=False, bgcolor='white', bordercolor=inh_colors['NA'], borderwidth=2,
                 type="buttons",
                 direction="left", font_size=10, font_family="Arial",
                 active=-1,
                 pad={"t": 10}, x=0.89, y=0.2,
                 buttons=[dict(args=[{}], label='NA', method='skip')])
        ])

    fig.update_layout(
        width=1000, height=750,
        autosize=False,
        margin=dict(t=100, b=200, l=100, r=100, pad=0),
    )

    fig.update_layout(
        annotations=[
            dict(text="DiGePred<br>score", x=0.039, xref="paper", y=-0.355, yref="paper", font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="gene or protein<br>interactions", x=0.173, xref="paper", y=-0.355, yref="paper", font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="KEGG and Reactome<br>pathway similarity", x=0.4, xref="paper", y=-0.355, yref="paper",
                 font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="HPO phenotype<br>similarity", x=0.6, xref="paper", y=-0.355, yref="paper", font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="co-expression<br>rank", x=0.85, xref="paper", y=-0.355, yref="paper", font_size=14,
                 font_family="Arial",
                 align="center", showarrow=False),
            dict(text="Select Data", x=1.02, xref="paper", y=-0.26, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="|", x=0.16, xref="paper", y=-0.33, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="|", x=0.295, xref="paper", y=-0.33, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="|", x=0.5, xref="paper", y=-0.33, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
            dict(text="|", x=0.71, xref="paper", y=-0.33, yref="paper", font_size=18,
                 font_family="Arial",
                 align="left", showarrow=False),
        ]
    )

    fig.update_layout(xaxis_showgrid=True, yaxis_showgrid=True,
                      title_text="DiGePred score | cutoff >= 0.1", hoverlabel=dict(
            bgcolor="white",
            font_size=16,
            font_family="Arial"
        ),
                      )
    fig.update_xaxes(tickson='boundaries', showgrid=True, gridcolor='black', gridwidth=2, zeroline=True,
                     showspikes=True, spikecolor='#74c476', spikethickness=2, spikedash='dot',
                     spikemode='toaxis+marker')
    fig.update_yaxes(tickson='boundaries', showgrid=True, gridcolor='black', gridwidth=2, zeroline=True,
                     showspikes=True, spikecolor='#74c476', spikethickness=2, spikedash='dot',
                     spikemode='toaxis+marker')

    fig['layout']['xaxis']['scaleanchor'] = 'y'
    fig['layout']['xaxis']['constrain'] = 'domain'
    fig['layout']['xaxis']['constraintoward'] = 'left'
    fig['layout']['xaxis']['scaleratio'] = 1

    pio.write_html(fig, file=html_file_name)
    print('html filename: {}'.format(os.path.abspath(html_file_name)))


def draw_inh_col_boxes_pair_level(fig, inh_df):
    for i, ga in enumerate(inh_df.index):
        for j, gb in enumerate(inh_df.columns):
            if ga != gb:
                ip = '{ga},{gb}'.format(ga=tuple(sorted([ga, gb]))[0],
                                        gb=tuple(sorted([ga, gb]))[1])
                if ip in dig_info_dict:
                    inh = dig_info_dict[ip]['inheritance']
                    zygA = dig_info_dict[ip]['variants'][ga]['zyg']
                    zygB = dig_info_dict[ip]['variants'][gb]['zyg']
                    inhA = dig_info_dict[ip]['variants'][ga]['inh']
                    inhB = dig_info_dict[ip]['variants'][gb]['inh']
                else:
                    inh = 'NA'
                    zygA = 'NA'
                    zygB = 'NA'
                    inhA = 'NA'
                    inhB = 'NA'

                inhzygA = inhA + str(zygA)
                inhzygB = inhB + str(zygB)

                if 'de novo' in inhzygA or 'de novo' in inhzygB:
                    fig.add_shape(type="rect",
                                  x0=j - 0.4, y0=i - 0.4, x1=j + 0.4, y1=i + 0.4,
                                  line=dict(
                                      color=inh_colors['de novo'],
                                      width=2,
                                  ),
                                  )
                elif 'homozygous' in inhzygA or 'homozygous' in inhzygB:
                    fig.add_shape(type="rect",
                                  x0=j - 0.4, y0=i - 0.4, x1=j + 0.4, y1=i + 0.4,
                                  line=dict(
                                      color=inh_colors['homozygous'],
                                      width=2,
                                  ),
                                  )
                elif 'compound het.' in inhzygA or 'compound het.' in inhzygB:
                    fig.add_shape(type="rect",
                                  x0=j - 0.4, y0=i - 0.4, x1=j + 0.4, y1=i + 0.4,
                                  line=dict(
                                      color=inh_colors['comp. het.'],
                                      width=2,
                                  ),
                                  )
                elif 'bi-parental' in inh:
                    fig.add_shape(type="rect",
                                  x0=j - 0.4, y0=i - 0.4, x1=j + 0.4, y1=i + 0.4,
                                  line=dict(
                                      color=inh_colors['bi-parental'],
                                      width=2,
                                  ),
                                  )
                elif 'X-linked' in inhzygA or 'X-linked' in inhzygB:
                    fig.add_shape(type="rect",
                                  x0=j - 0.4, y0=i - 0.4, x1=j + 0.4, y1=i + 0.4,
                                  line=dict(
                                      color=inh_colors['X-linked'],
                                      width=2,
                                  ),
                                  )
                elif 'maternal' in inhzygA and 'paternal' not in inhzygA and \
                        'maternal' in inhzygB and 'paternal' not in inhzygB:
                    fig.add_shape(type="rect",
                                  x0=j - 0.4, y0=i - 0.4, x1=j + 0.4, y1=i + 0.4,
                                  line=dict(
                                      color=inh_colors['maternal'],
                                      width=2,
                                  ),
                                  )
                elif 'paternal' in inhzygA and 'maternal' not in inhzygA and \
                        'paternal' in inhzygB and 'maternal' not in inhzygB:
                    fig.add_shape(type="rect",
                                  x0=j - 0.4, y0=i - 0.4, x1=j + 0.4, y1=i + 0.4,
                                  line=dict(
                                      color=inh_colors['paternal'],
                                      width=2,
                                  ),
                                  )
                else:
                    fig.add_shape(type="rect",
                                  x0=j - 0.4, y0=i - 0.4, x1=j + 0.4, y1=i + 0.4,
                                  line=dict(
                                      color=inh_colors['NA'],
                                      width=2,
                                  ),
                                  )

    return fig


def make_json(df):
    dig_info_dict = dict()
    genes = sorted(list(set(df['gene A']).union(df['gene B'])))
    gene_info_dict = {}
    for g in genes:
        gene_info_dict[g] = {}
        gene_info_dict[g]['inh'] = list(set(list(dig_df[(dig_df['gene A'] == g)]['inheritance A'].values) +
                                            list(dig_df[(dig_df['gene B'] == g)]['inheritance B'].values)))[0]

        gene_info_dict[g]['zyg'] = list(set(list(dig_df[(dig_df['gene A'] == g)]['zygosity A'].values) +
                                            list(dig_df[(dig_df['gene B'] == g)]['zygosity B'].values)))[0]

        gene_info_dict[g]['consq'] = list(set(list(dig_df[(dig_df['gene A'] == g)]['consq. A'].values) +
                                              list(dig_df[(dig_df['gene B'] == g)]['consq. B'].values)))[0]

        gene_info_dict[g]['cdnavar'] = list(set(list(dig_df[(dig_df['gene A'] == g)]['cDNA change A'].values) +
                                                list(dig_df[(dig_df['gene B'] == g)]['cDNA change B'].values)))[0]

        gene_info_dict[g]['protvar'] = list(set(list(dig_df[(dig_df['gene A'] == g)]['protein change A'].values) +
                                                list(dig_df[(dig_df['gene B'] == g)]['protein change B'].values)))[0]

    for i in df.index:
        ga = df['gene A'][i]
        gb = df['gene B'][i]
        score = '%0.3g' % float(df['digenic score'][i])
        pcile = int(round(float(df['percentile'][i]), 0))
        if pcile == 100:
            pcile = 99
        if pcile == 0:
            pcile = 1

        dir_int = [1 if df['PPI distance'][i] == 1 or
                        df['pathway distance'][i] == 1 or
                        df['literature distance'][i] == 1 else 'NA'][0]
        indir_int = [1 if df['PPI distance'][i] == 2 or
                          df['pathway distance'][i] == 2 or
                          df['literature distance'][i] == 2 else np.nan][0]
        coex = [df['mutual co-expression rank'][i] if
                df['mutual co-expression rank'][i] > 0 else 'NA'][0]

        inh = df['pair inheritance'][i]
        pwy = np.round(df['pathway similarity'][i], 2)
        pheno = np.round(df['phenotype similarity'][i], 2)

        row = ga
        col = gb

        if row != col:

            ip = '{ga},{gb}'.format(ga=tuple(sorted([row, col]))[0],
                                    gb=tuple(sorted([row, col]))[1])

            if ip not in dig_info_dict:
                dig_info_dict[ip] = {'inheritance': '',
                                     'variants': {ga: {'cDNA': gene_info_dict[ga]['cdnavar'][:-1],
                                                       'prot': gene_info_dict[ga]['protvar'][:-1],
                                                       'consq': gene_info_dict[ga]['consq'][:-1],
                                                       'inh': gene_info_dict[ga]['inh'],
                                                       'zyg': gene_info_dict[ga]['zyg'][:-1]},
                                                  gb: {'cDNA': gene_info_dict[gb]['cdnavar'][:-1],
                                                       'prot': gene_info_dict[gb]['protvar'][:-1],
                                                       'consq': gene_info_dict[gb]['consq'][:-1],
                                                       'inh': gene_info_dict[gb]['inh'],
                                                       'zyg': gene_info_dict[gb]['zyg'][:-1]},
                                                  },
                                     'distance': 0,
                                     'path': [],
                                     'interactions': {},
                                     'pathways': {'similarity': 0, 'common': {'codes': [], 'names': []},
                                                  '# common': 0, '# total': 0
                                                  },
                                     'phenotypes': {'similarity': 0, 'common': {'codes': [], 'names': []},
                                                    '# common': 0, '# total': 0
                                                    },
                                     'coexpression': {'rank': 0, 'tissues': []},
                                     'DiGePred': {'score': score, 'percentile': pcile}}

            if inh != "NA":
                if inh == 'both parents not sequenced':
                    dig_info_dict[ip]['inheritance'] = 'NA'
                elif 'de novo' not in inh:
                    dig_info_dict[ip]['inheritance'] = '{inh} inheritance'.format(inh=inh)
                else:
                    dig_info_dict[ip]['inheritance'] = 'de novo'
            else:
                dig_info_dict[ip]['inheritance'] = 'NA'

            if dir_int == 1:
                dig_info_dict[ip]['distance'] = 1
                dig_info_dict[ip]['path'] = sorted(ip.split(','))
                int_details_dict = get_int_details(geneA=row, geneB=col)
                dig_info_dict[ip]['interactions'][ip] = {'types': int_details_dict['types'],
                                                         'sources': int_details_dict['sources'],
                                                         'evidences': list(set(int_details_dict['evidences'])),
                                                         'references': int_details_dict['references'],
                                                         }

            elif indir_int == 1:
                path, path_details_dict = get_shortest_path_details(row, col)
                if path:
                    if len(path) > 2:
                        dig_info_dict[ip]['distance'] = 2
                        dig_info_dict[ip]['path'] = path
                        for gp in path_details_dict:
                            gx = '{ga},{gb}'.format(ga=sorted(list(gp))[0],
                                                    gb=sorted(list(gp))[1])
                            dig_info_dict[ip]['interactions'][gx] = {'types': list(set(path_details_dict[gp]['types'])),
                                                                     'sources': list(
                                                                         set(path_details_dict[gp]['sources'])),
                                                                     'evidences': list(
                                                                         set(path_details_dict[gp]['evidences'])),
                                                                     'references': path_details_dict[gp]['references'],
                                                                     }
                else:
                    dig_info_dict[ip]['distance'] = '>2'
                    dig_info_dict[ip]['path'] = []

            if row in kegg_gene_to_codes_dict:
                row_pwys = kegg_gene_to_codes_dict[row]
            else:
                row_pwys = []
            if row in reactome_gene_to_codes_dict:
                row_pwys.extend(reactome_gene_to_codes_dict[row])

            if col in kegg_gene_to_codes_dict:
                col_pwys = kegg_gene_to_codes_dict[col]
            else:
                col_pwys = []
            if col in reactome_gene_to_codes_dict:
                col_pwys.extend(reactome_gene_to_codes_dict[col])

            if row_pwys and col_pwys:
                com_pwys = list(set(row_pwys).intersection(col_pwys))

                com_pwy_names = sorted(
                    set([kegg_code_to_name_dict[c] for c in com_pwys if
                         c in kegg_code_to_name_dict] +
                        [reactome_code_to_name_dict[c] for c in com_pwys if
                         c in reactome_code_to_name_dict]))

                dig_info_dict[ip]['pathways']['similarity'] = pwy
                dig_info_dict[ip]['pathways']['# common'] = len(com_pwys)
                dig_info_dict[ip]['pathways']['# total'] = len(set(row_pwys).union(col_pwys))
                dig_info_dict[ip]['pathways']['common'] = {'codes': [], 'names': []}

                dig_info_dict[ip]['pathways']['common']['codes'] = com_pwys
                dig_info_dict[ip]['pathways']['common']['names'] = com_pwy_names

            if row in pheno_gene_to_codes_dict:
                row_phenos = pheno_gene_to_codes_dict[row]
            else:
                row_phenos = []
            if col in pheno_gene_to_codes_dict:
                col_phenos = pheno_gene_to_codes_dict[col]
            else:
                col_phenos = []

            if row_phenos and col_phenos:
                com_phenos = list(set(row_phenos).intersection(col_phenos))

                dig_info_dict[ip]['phenotypes']['similarity'] = pheno
                dig_info_dict[ip]['phenotypes']['# common'] = len(com_phenos)
                dig_info_dict[ip]['phenotypes']['# total'] = len(set(row_phenos).union(col_phenos))
                dig_info_dict[ip]['phenotypes']['common'] = {'codes': [], 'names': []}
                pheno_names = sorted(set([pheno_code_to_name_dict[c]
                                          for c in set(com_phenos) if
                                          c in pheno_code_to_name_dict and
                                          'obsolete' not in pheno_code_to_name_dict[
                                              c].lower()]))

                dig_info_dict[ip]['phenotypes']['common']['codes'] = com_phenos
                dig_info_dict[ip]['phenotypes']['common']['names'] = pheno_names

            if row in gene_to_mean_TPM_by_tissue_dict:
                tpmsa = np.asarray(gene_to_mean_TPM_by_tissue_dict[row])
                tpmsb = []
                if np.sum(tpmsa) > 0:
                    tpmsa = tpmsa / np.sum(tpmsa)
                    if col in gene_to_mean_TPM_by_tissue_dict:
                        tpmsb = np.asarray(gene_to_mean_TPM_by_tissue_dict[col])
                        if np.sum(tpmsb) > 0:
                            tpmsb = tpmsb / np.sum(tpmsb)

                    highest_ex_tissues_a = []
                    highest_ex_tissues_b = []
                    tpm_vals_dict = dict()

                    for i, vv in sorted([(i, x) for (i, x) in enumerate(tpmsa)],
                                        key=lambda x: x[1], reverse=True):
                        if vv > 0:
                            highest_ex_tissues_a.append(id_to_narrow_tissue_dict[str(i)])

                        tpm_vals_dict[id_to_narrow_tissue_dict[str(i)]] = vv

                    for i, vv in sorted([(i, x) for (i, x) in enumerate(tpmsb)],
                                        key=lambda x: x[1], reverse=True):
                        if vv > 0:
                            highest_ex_tissues_b.append(id_to_narrow_tissue_dict[str(i)])

                        tpm_vals_dict[id_to_narrow_tissue_dict[str(i)]] += vv

                    highest_ex_tissues_com = set(highest_ex_tissues_a[:10]).intersection(
                        highest_ex_tissues_b[:10])
                    highest_ex_tissues_com = sorted(highest_ex_tissues_com,
                                                    key=lambda x: tpm_vals_dict[x],
                                                    reverse=True)

            dig_info_dict[ip]['coexpression']['rank'] = [round(coex, 0) if coex != 'NA' else 'NA'][0]
            if coex != 'NA':
                dig_info_dict[ip]['coexpression']['tissues'] = highest_ex_tissues_com

    json.dump(dig_info_dict, open(details_json_name, 'w'), indent=6)
    print('summary dig_info_dict filename: {}'.format(os.path.abspath(details_json_name)))

    return dig_info_dict


def get_inh_dict(dig_info_dict):
    gp_inh_dict = dict()
    inh_gp_dict = dict()
    for ip in dig_info_dict:
        ga, gb = tuple(sorted(ip.split(',')))
        ds = dig_info_dict[ip]['DiGePred']['score']
        dp = dig_info_dict[ip]['DiGePred']['percentile']
        path = dig_info_dict[ip]['path']
        dist = dig_info_dict[ip]['distance']
        inh = dig_info_dict[ip]['inheritance']
        zygA = dig_info_dict[ip]['variants'][ga]['zyg']
        zygB = dig_info_dict[ip]['variants'][gb]['zyg']
        inhA = dig_info_dict[ip]['variants'][ga]['inh']
        inhB = dig_info_dict[ip]['variants'][gb]['inh']

        inhzygA = inhA + str(zygA)
        inhzygB = inhB + str(zygB)

        iza = ' '
        if 'de novo' in inhzygA:
            iza += '<i>de novo</i>,'
        if 'homozygous' in inhzygA:
            iza += 'homozygous,'
        if 'comp het' in inhzygA:
            iza += 'compound het.,'
        if 'X-linked' in inhzygA:
            iza += 'X-linked,'
        if 'maternal' in inhzygA and 'paternal' not in inhzygA:
            iza += 'maternal,'
        if 'paternal' in inhzygA and 'maternal' not in inhzygA:
            iza += 'paternal,'

        izb = ' '
        if 'de novo' in inhzygB:
            izb += '<i>de novo</i>,'
        if 'homozygous' in inhzygB:
            izb += 'homozygous,'
        if 'comp het' in inhzygB:
            izb += 'compound het.,'
        if 'X-linked' in inhzygB:
            izb += 'X-linked,'
        if 'maternal' in inhzygB and 'paternal' not in inhzygB:
            izb += 'maternal,'
        if 'paternal' in inhzygB and 'maternal' not in inhzygB:
            izb += 'paternal,'

        if 'de novo' in iza or 'de novo' in izb:
            iz = 'de novo'
            u = 'unique in proband'
        elif 'homozygous' in iza or 'homozygous' in izb:
            iz = 'homozygous'
            u = 'unique in proband'
        elif 'compound het.' in iza or 'compound het.' in izb:
            iz = 'compound het.'
            u = 'unique in proband'
        elif 'maternal' in iza and 'paternal' in izb:
            iz = 'bi-parental'
            u = 'unique in proband'
        elif 'paternal' in iza and 'maternal' in izb:
            iz = 'bi-parental'
            u = 'unique in proband'
        elif 'X-linked' in iza or 'X-linked' in izb:
            iz = 'X-linked'
            u = 'unique in proband'
        elif 'maternal' in iza and 'maternal' in izb:
            iz = 'maternal'
            u = ''
        elif 'paternal' in iza and 'paternal' in izb:
            iz = 'paternal'
            u = ''
        else:
            iz = 'NA'
            u = ''

        iza = iza[1:-1]
        izb = izb[1:-1]

        if iz not in inh_gp_dict:
            inh_gp_dict[iz] = []
        inh_gp_dict[iz].append((ga, gb))

        gp_inh_dict[(ga, gb)] = (iz, iza, izb, u)
        gp_inh_dict[(gb, ga)] = (iz, izb, iza, u)

    return gp_inh_dict, inh_gp_dict

def clean_consequence(consequence) -> str:
    consequence = consequence.lower().replace('variant', '')
    if '(' in consequence:
        consequence = consequence[: consequence.find('(')]
    consequence = consequence.lower().strip().rstrip()
    if 'splice' in consequence:
        consequence = 'splicing'


    return consequence



# Dead function July 10 2024
def parse_excel_file(excel_file_path):
    parsed_excel_file = get_parsed_excel_file(excel_file_path)
    relcols_dict, rels, label_row = get_inh_cols(parsed_excel_file)

    genes = []
    tscripts = []
    for rnum, row in enumerate(parsed_excel_file):
        if row[0] is not None:
            inh, zyg = get_inheritance_zygosity(row, relcols_dict, rels)
            g, tscript = get_gene_tscript(row)
            if inh and g.lower() != 'gene' and ':' not in g:
                gs = [x.strip().rstrip() for x in g.split(',')]
                for gene in gs:
                    chr, pos, rsid, ref, alt, consequence = get_chr_coord(parsed_excel_file, row, rnum)
                    protvar = get_prot_var(parsed_excel_file, row, rnum)
                    cdnavar = get_cdna_var(parsed_excel_file, row, rnum)

                    consequence = clean_consequence(consequence)

                    disp_name, ref_name = match_gene_tscript_names(gene, tscript)
                    disp_name, ref_name = check_gene_alias(ref_name)

                    gene_disp_name_to_ref_name_dict[disp_name] = ref_name
                    gene_ref_name_to_disp_name_dict[ref_name] = disp_name

                    genes.append(disp_name)
                    if tscript != '':
                        tscripts.append(tscript)

                    if disp_name not in gene_var_details_dict:
                        gene_var_details_dict[disp_name] = {'tscript': tscript,
                                                            'chr': '',
                                                            'pos': '',
                                                            'ref': '',
                                                            'alt': '',
                                                            'rs ID': '',
                                                            'consequence': '',
                                                            'cdna var': '',
                                                            'prot var': '',
                                                            'zyg': '',
                                                            'inh': ''
                                                            }

                    gene_var_details_dict[disp_name]['chr'] = str(chr)
                    gene_var_details_dict[disp_name]['pos'] += str(pos) + ';'
                    gene_var_details_dict[disp_name]['ref'] += ref + ';'
                    gene_var_details_dict[disp_name]['alt'] += alt + ';'
                    gene_var_details_dict[disp_name]['rs ID'] += str(rsid) + ';'
                    gene_var_details_dict[disp_name]['consequence'] += consequence + ';'
                    gene_var_details_dict[disp_name]['cdna var'] += cdnavar + ';'
                    gene_var_details_dict[disp_name]['prot var'] += protvar + ';'
                    gene_var_details_dict[disp_name]['zyg'] += zyg + ';'
                    gene_var_details_dict[disp_name]['inh'] += inh + ';'

    pairs = itertools.combinations(set(genes), 2)

    return gene_var_details_dict, pairs


vtypes = ['stop gained', 'missense', 'deletion', 'duplication', 'frameshift', 'splicing', 'intronic', 'other',
          'synonymous']

if args.vustruct:
    # 2024 July.  Chris is using the generic _vustruct.csv format to provide all inheritance details
    # We keep the UDN excel format (args.excel) as a way to compare this new parser results to existing one
    # I.e. until we are sure we are doing vustruct right, we keep comparing to original spreadsheet when we have it
    genes = []
    tscripts = []
    vustruct_filename = args.vustruct
    vustruct_case_df = pd.read_csv(vustruct_filename, sep=',', index_col=None,
                                       keep_default_na=False, encoding='utf8',
                                       comment='#', skipinitialspace=True)
    rels = [] # a list of strings which might include mother, father, etc  Maybe more - need to think
    if 'inheritance' in vustruct_case_df.columns:
        rel_set = set()
        for index, row in vustruct_case_df.iterrows():
            for possible_rel in ['mother', 'father']:
                if possible_rel in row['inheritance']:
                    rel_set.add(possible_rel)
        rels = sorted(rel_set) # So we have now a list like ['mother', 'father'] if the other familiy members were sequenced

    gene_already_seen = set()

    for index, row in vustruct_case_df.iterrows():
        gene = row.get('gene','')
        g = gene
        if not gene:
            LOGGER.error("No gene entry found in row %s.  Skipping....", str(row))
        chr = row.get('chrom','')
        pos = row.get('pos','')
        ref = row.get('ref','')
        alt = row.get('alt','')

        gene_tuple=(gene,chr,pos,ref,alt)
        if gene_tuple in gene_already_seen:
            continue

        gene_already_seen.add(gene_tuple)

        tscript = row.get('refseq', '')

        consequence = clean_consequence(row.get('effect',''))

        disp_name, ref_name = match_gene_tscript_names(gene, tscript)
        disp_name, ref_name = check_gene_alias(ref_name)

        gene_disp_name_to_ref_name_dict[disp_name] = ref_name
        gene_ref_name_to_disp_name_dict[ref_name] = disp_name

        inh = row.get('inheritance','')
        if not inh:
            inh = "NA" # A Shourid notation that seems important
        zyg = row.get('zygosity','')
        rsid = ''
        protvar = row.get('mutation','')
        cdnavar = ''
        genes.append(disp_name)
        if tscript != '':
            tscripts.append(tscript)

        if disp_name not in gene_var_details_dict:
            gene_var_details_dict[disp_name] = {'tscript': tscript,
                                                'chr': '',
                                                'pos': '',
                                                'ref': '',
                                                'alt': '',
                                                'rs ID': '',
                                                'consequence': '',
                                                'cdna var': '',
                                                'prot var': '',
                                                'zyg': '',
                                                'inh': ''}
 
        gene_var_details_dict[disp_name]['chr'] = str(chr)
        gene_var_details_dict[disp_name]['pos'] += str(pos) + ';'
        gene_var_details_dict[disp_name]['ref'] += ref + ';'
        gene_var_details_dict[disp_name]['alt'] += alt + ';'
        gene_var_details_dict[disp_name]['rs ID'] += str(rsid) + ';'
        gene_var_details_dict[disp_name]['consequence'] += consequence + ';'
        gene_var_details_dict[disp_name]['cdna var'] += cdnavar + ';'
        gene_var_details_dict[disp_name]['prot var'] += protvar + ';'
        gene_var_details_dict[disp_name]['zyg'] += zyg + ';'
        gene_var_details_dict[disp_name]['inh'] += inh + ';'

        
    pairs = itertools.combinations(set(genes), 2)

elif args.excel:
    excel_file_name = args.excel
    parsed_excel_file = get_parsed_excel_file(excel_file_name)
    # Parse out the available inheritance information
    relcols_dict, rels, label_row = get_inh_cols(parsed_excel_file)

    genes = []
    tscripts = []
    for rnum, row in enumerate(parsed_excel_file):
        if row[0] is not None:
            inh, zyg = get_inheritance_zygosity(row, relcols_dict, rels)
            g, tscript = get_gene_tscript(row)
            if inh and g.lower() != 'gene' and ':' not in g:
                gs = [x.strip().rstrip() for x in g.split(',')]

                ## Chris moth needs to figure out why Souhrid is parsing multi genes here

                for gene in gs:
                    chr, pos, rsid, ref, alt, consequence = get_chr_coord(parsed_excel_file, row, rnum)
                    protvar = get_prot_var(parsed_excel_file, row, rnum)
                    cdnavar = get_cdna_var(parsed_excel_file, row, rnum)

                    consequence = clean_consequence(consequence)

                    disp_name, ref_name = match_gene_tscript_names(gene, tscript)
                    disp_name, ref_name = check_gene_alias(ref_name)

                    gene_disp_name_to_ref_name_dict[disp_name] = ref_name
                    gene_ref_name_to_disp_name_dict[ref_name] = disp_name

                    genes.append(disp_name)
                    if tscript != '':
                        tscripts.append(tscript)

                    if disp_name not in gene_var_details_dict:
                        gene_var_details_dict[disp_name] = {'tscript': tscript,
                                                            'chr': '',
                                                            'pos': '',
                                                            'ref': '',
                                                            'alt': '',
                                                            'rs ID': '',
                                                            'consequence': '',
                                                            'cdna var': '',
                                                            'prot var': '',
                                                            'zyg': '',
                                                            'inh': ''
                                                            }

                    gene_var_details_dict[disp_name]['chr'] = str(chr)
                    gene_var_details_dict[disp_name]['pos'] += str(pos) + ';'
                    gene_var_details_dict[disp_name]['ref'] += ref + ';'
                    gene_var_details_dict[disp_name]['alt'] += alt + ';'
                    gene_var_details_dict[disp_name]['rs ID'] += str(rsid) + ';'
                    gene_var_details_dict[disp_name]['consequence'] += consequence + ';'
                    gene_var_details_dict[disp_name]['cdna var'] += cdnavar + ';'
                    gene_var_details_dict[disp_name]['prot var'] += protvar + ';'
                    gene_var_details_dict[disp_name]['zyg'] += zyg + ';'
                    gene_var_details_dict[disp_name]['inh'] += inh + ';'

    
    pairs = itertools.combinations(set(genes), 2)
    

elif args.csv:
    print('Found gene transcript variant zygosity inheritance csv / tsv file')
    gene_var_inh_file = open(args.csv).read().split('\n')[:-1]
    genes = []
    tscripts = []

    for line in gene_var_inh_file:
        if ',' in line:
            x = line.split(',')
        elif '\t' in line:
            x = line.split('\t')
        else:
            print('gene variant inheritance file not csv / tsv')
            break

        gene = x[0].strip().rstrip()
        tscript = x[1].strip().rstrip()
        cv = x[2].strip().rstrip()
        pv = x[3].strip().rstrip()
        z = x[4].strip().rstrip()
        inh = x[5].strip().rstrip()

        disp_name, ref_name = match_gene_tscript_names(gene, tscript)
        disp_name, ref_name = check_gene_alias(ref_name)

        gene_disp_name_to_ref_name_dict[disp_name] = ref_name
        gene_ref_name_to_disp_name_dict[ref_name] = disp_name

        genes.append(disp_name)
        if t != '':
            tscripts.append(tscript)

        if disp_name not in gene_var_details_dict:
            gene_var_details_dict[disp_name] = {'tscript': t,
                                                'cdna var': [],
                                                'prot var': [],
                                                'zyg': [],
                                                'inh': []
                                                }

        gene_var_details_dict[disp_name]['cdna var'].append(cv)
        gene_var_details_dict[disp_name]['prot var'].append(pv)
        gene_var_details_dict[disp_name]['zyg'].append(z)
        gene_var_details_dict[disp_name]['inh'].append(inh)

    pairs = itertools.combinations(set(genes), 2)

elif args.pairs:
    print('Found user input gene pairs file')
    pairs_list = open(args.pairs).read().split('\n')[:-1]
    tscript = ''
    tscripts = []

    for line in pairs_list:
        g1 = line.split(',')[0].strip().rstrip()
        g2 = line.split(',')[1].strip().rstrip()

        disp_name1, ref_name1 = match_gene_tscript_names(g1, tscript)
        disp_name1, ref_name1 = check_gene_alias(ref_name1)

        gene_disp_name_to_ref_name_dict[disp_name1] = ref_name1
        gene_ref_name_to_disp_name_dict[ref_name1] = disp_name1

        disp_name2, ref_name2 = match_gene_tscript_names(g2, tscript)
        disp_name2, ref_name2 = check_gene_alias(ref_name2)

        gene_disp_name_to_ref_name_dict[disp_name2] = ref_name2
        gene_ref_name_to_disp_name_dict[ref_name2] = disp_name2

        genes.append(disp_name1)
        genes.append(disp_name2)

        if disp_name1 not in gene_var_details_dict:
            gene_var_details_dict[disp_name1] = {'tscript': tscript,
                                                 'chr': '',
                                                 'pos': '',
                                                 'ref': '',
                                                 'alt': '',
                                                 'rs ID': '',
                                                 'consequence': '',
                                                 'cdna var': '',
                                                 'prot var': '',
                                                 'zyg': '',
                                                 'inh': ''
                                                 }

        if disp_name2 not in gene_var_details_dict:
            gene_var_details_dict[disp_name2] = {'tscript': tscript,
                                                 'chr': '',
                                                 'pos': '',
                                                 'ref': '',
                                                 'alt': '',
                                                 'rs ID': '',
                                                 'consequence': '',
                                                 'cdna var': '',
                                                 'prot var': '',
                                                 'zyg': '',
                                                 'inh': ''
                                                 }

        pairs.append(tuple(sorted([ref_name1, ref_name2])))

    pairs = sorted(set(pairs))

elif args.genes:
    print('Found user input genes file')

    genes = []
    tscripts = []
    tscript = ''

    for g in sorted(set(open(args.genes).read().split('\n')[:-1])):
        if g != '':
            disp_name, ref_name = match_gene_tscript_names(g, tscript)
            disp_name, ref_name = check_gene_alias(ref_name)

            gene_disp_name_to_ref_name_dict[disp_name] = ref_name
            gene_ref_name_to_disp_name_dict[ref_name] = disp_name

            genes.append(disp_name)

            if disp_name not in gene_var_details_dict:
                gene_var_details_dict[disp_name] = {'tscript': tscript,
                                                    'chr': '',
                                                    'pos': '',
                                                    'ref': '',
                                                    'alt': '',
                                                    'rs ID': '',
                                                    'consequence': '',
                                                    'cdna var': '',
                                                    'prot var': '',
                                                    'zyg': '',
                                                    'inh': 'NA'
                                                    }

    pairs = itertools.combinations(set(genes), 2)

else:
    print('No input file found!')

dig_df, sel_gp_csv_df = get_digenic_metrics_csv(gene_pairs=pairs)
dig_info_dict = make_json(dig_df)
if len(dig_df.index) < 500:
    plot_html(dig_df, dig_info_dict)
