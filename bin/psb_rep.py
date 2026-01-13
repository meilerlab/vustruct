#!/usr/bin/env python3
#
# Project        : PSB Pipeline
# Filename       : psb_rep.py
# Authors        : Chris Moth and R. Michael Sivley
# project Organization   : Vanderbilt Genetics Institute,
#                : Vanderbilt University
# Email          : chris.moth@vanderbilt.edu
# Date           : 2018-01-22
# Description    : Output final html and pdf reports from completed jobs
#                : (output from psb_launch.py)
# =============================================================================#


"""\
Create final html and pdf reports for completed pipeline jobs from either
a project id (ex. UDN123456) or a single workstatus.csv file output from psb_monitor.py

Configuration options must be provided with
   -c global.config file
   -u user.config overrides
   structure_report.csv originally output by psb_plan.py
   workstatus.csv as previously updated with psb_monitor.py

HTML is created by populating template .html through jinja2 template library
"""

import datetime
import grp
import logging
import os
# import argparse, configparser
import pprint
import string
import math
import sys
import time
import html
from array import array
from logging.handlers import RotatingFileHandler

import shutil
import pandas as pd
from jinja2 import Environment, FileSystemLoader
from io import StringIO
import json

import traceback

def global_psb_rep_exception_handler(exc_type, exc_value, exc_traceback):
    """If there is a major problem dring psb_rep report generation, we need to log that exception as the web page"""
    logging.getLogger().info("Fail of psb_rep.py.  Exception: %s: %s", exc_type.__name__, exc_value)
    traceback.print_exception(exc_type, exc_value, exc_traceback, file=sys.stderr)

sys.excepthook = global_psb_rep_exception_handler

# openpyxl supports creation of a case summary of summary
# worksheet
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
# from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

# matplotlib for creation of separate .png file showing
# ScanNetPPIs and MusiteDeepPTM prediction results
# per Alican's manual calculations
import matplotlib.pyplot as plt

import subprocess
from lib import PDBMapSQLdb
from lib import PDBMapTranscriptBase
from lib import PDBMapTranscriptUniprot
from lib import PDBMapTranscriptEnsembl
from lib import PDBMapProtein
from lib import PDBMapGlobals

from vustruct import VUStruct
from report_modules.rep_one_variant import report_one_variant_one_isoform
# from report_modules.load_one_variant import CalculationResultsLoader



from collections import defaultdict

from typing import Dict, List
# from subprocess import Popen, PIPE
# from weasyprint import HTML

from psb_shared import psb_config

print("4 of 4  %s Pipeline report generator.  -h for detailed help" % __file__)

sh = logging.StreamHandler()
LOGGER = logging.getLogger()
LOGGER.addHandler(sh)

# Now that we've added streamHandler, basicConfig will not add another handler (important!)
log_format_string = '%(asctime)s %(levelname)-4s [%(filename)16s:%(lineno)d] %(message)s'
date_format_string = '%H:%M:%S'
log_formatter = logging.Formatter('%(asctime)s %(levelname)-4s [%(filename)20s:%(lineno)d] %(message)s',
                                  datefmt="%H:%M:%S")

LOGGER.setLevel(logging.DEBUG)
sh.setLevel(logging.INFO)
sh.setFormatter(log_formatter)

root_dir_log_filename = "psb_rep.log"
need_roll = os.path.isfile(root_dir_log_filename)
rootdir_fh = RotatingFileHandler(root_dir_log_filename, backupCount=7)

rootdir_fh.setFormatter(log_formatter)
rootdir_fh.setLevel(logging.INFO)
LOGGER.addHandler(rootdir_fh)

if need_roll:
    rootdir_fh.doRollover()

sys.stderr.write("Log file for entire case is %s\n" % root_dir_log_filename)

# As report is generated, build up a list of just the files needed to make a portable website
# This list omits the intermediate files left behind in the calculation
website_filelist = ['html/']

cmdline_parser = psb_config.create_default_argument_parser(__doc__, os.path.dirname(os.path.dirname(__file__)))

cmdline_parser.add_argument("-s", "--slurm",
                            help="Create a slurm file to launch all the reports", action="store_true")
cmdline_parser.add_argument("--embed_refresh",
                            help="Set if created pages should auto refresh, i.e. are not final", action="store_true")
cmdline_parser.add_argument("--refresh_interval_seconds", nargs='?', type=int, metavar='int', 
                            help="Set the interval in which the page should reload itself", default=0)
cmdline_parser.add_argument("--seconds_remaining", nargs='?', type=int, metavar='int', 
                            help="Add text to let user know how much longer report re-generation will continue", default=0)
cmdline_parser.add_argument("--user_message", type=str, nargs='?',
                            help="Include a message to the user in the landing page, such as 'report generation complete'")
cmdline_parser.add_argument("--tar_only",
                            help="Do NOT create a .zip output of the website, just .tar", action="store_true")
cmdline_parser.add_argument("--strip_uuid", type=str, nargs='?',
                            help="If supplied, a uuid to remove from archived filenames (used after public website input")
cmdline_parser.add_argument("--web_case_id", type=str, nargs='?',
                            help="Case ID as input through user form")
cmdline_parser.add_argument("projectORstructure", type=str,
                            help="The project ID UDN123456 to report on all mutations.  Else, a specific structures file from psb_plan.py  Example: ....../UDN/UDN123456/GeneName_NM_12345.1_S123A_structure_report.csv",
                            default=os.path.basename(os.getcwd()), nargs='?')
cmdline_parser.add_argument("workstatus", nargs='?', type=str,
                            help="Blank for all mutations, or specific output file from psb_monitor.py\n" +
                                 "Example: ....../UDN/UDN123456/GeneName_NM_12345.1_S123A_workstatus.csv")
args, remaining_argv = cmdline_parser.parse_known_args()

# A period in the argument means the user wants to monitor one mutation only,
# directly from a single mutation output file of psb_launch.py
oneMutationOnly = (
        '.' in args.projectORstructure and os.path.isfile(args.projectORstructure) and args.workstatus is not None)
infoLogging = False

if args.debug:
    infoLogging = True
    sh.setLevel(logging.DEBUG)
    rootdir_fh.setLevel(logging.DEBUG)
elif args.verbose:
    infoLogging = True
    sh.setLevel(logging.INFO)

# Load the structure details of all jobs that were not dropped

# print "Command Line Arguments"
LOGGER.info("Command Line Arguments:\n%s" % pprint.pformat(vars(args)))

if not os.path.exists(args.config):
    LOGGER.critical("Global config file not found: " + args.config)
    sys.exit(1)

required_config_items = ['output_rootdir', 'collaboration', 'ddg_config', 'rate4site_dir', 'cosmis_dir', 'chgrp_to']

config, config_dict = psb_config.read_config_files(args, required_config_items)
config_dict_shroud_password = {x: config_dict[x] for x in required_config_items}
dbpass = config_dict.get('dbpass', '?')
config_dict_shroud_password['dbpass'] = '*' * len(dbpass)
LOGGER.info("Configuration File parameters:\n%s" % pprint.pformat(config_dict_shroud_password))

LOGGER.info("Loading idmapping")
PDBMapProtein.load_idmapping(config_dict['idmapping'])
LOGGER.info("Loading done")


try:
    slurmParametersAll = dict(config.items("SlurmParametersAll"))
except Exception as ex:
    slurmParametersAll = {}

import copy

slurmParametersReport = copy.deepcopy(slurmParametersAll)

try:
    slurmParametersReport.update(dict(config.items("SlurmParametersReport")))
except Exception as ex:
    pass

config_pathprox_dict = dict(config.items("PathProx"))
LOGGER.info("Pathprox config entries:\n%s" % pprint.pformat(config_pathprox_dict))

# The case_root_dir is the master directory for the case, i.e. for one patient
# Example: /dors/capra_lab/projects/psb_collab/UDN/UDN532183
udn_root_directory = os.path.join(config_dict['output_rootdir'], config_dict['collaboration'])

if oneMutationOnly:
    collaboration_absolute_dir = os.path.dirname(os.path.dirname(args.projectORstructure))
else:
    collaboration_absolute_dir = os.path.join(udn_root_directory, args.projectORstructure)
    vustruct = VUStruct("report", args.projectORstructure, __file__)

collaboration_absolute_dir = os.path.realpath(collaboration_absolute_dir)
LOGGER.info("Collaboration_absolute_dir = %s", collaboration_absolute_dir)

case_root_dir = os.path.relpath(os.getcwd(), collaboration_absolute_dir)
LOGGER.info("Relative to cwd=%s, case_root_directory= %s", os.getcwd(), case_root_dir)

LOG_FILES_DIRECTORY="./log"

# This module will create a summary .xlsx,and some graphics.
# Might as well go ahead and create thier directory
REPORT_FILES_DIRECTORY="./report_files"
os.makedirs(REPORT_FILES_DIRECTORY, exist_ok=True)

def copy_html_css_javascript():
    # The source html directory, and it's large set of supporting files, is located beneath the directory
    # where this file, psb_rep.py, is located.  All mutations share this one large html/css/javascript
    # sourcee directory
    src = os.path.join(os.path.dirname(os.path.realpath(__file__)), "html")
    dest = "%s/html" % case_root_dir

    nglviewer_filename = os.path.join(dest, "nglviewer", "examples", "dist", "ngl.js")
    if os.path.exists(nglviewer_filename):
        LOGGER.warning("%s found.  Skipping copy_html_css_javascript", nglviewer_filename)
        return
    else:
        LOGGER.info("%s not found.  Assuming first time copy", nglviewer_filename)

    LOGGER.info("Copying supporting javascript and html from %s to %s" % (src, dest))

    try:
        shutil.rmtree(dest, ignore_errors=True)
        # Save some space by not coping the data from nglviewer
        shutil.copytree(src, dest, ignore=shutil.ignore_patterns('*/nglview/data/*'))
    except Exception as e:
        LOGGER.exception(e)
        sys.exit(1)

    # Make sure
    src_file_list = [os.path.join(root, name) for root, dirs, files in os.walk(src) for name in files]

    dest_file_list = [os.path.join(root, name) for root, dirs, files in os.walk(dest) for name in files]
    if len(src_file_list) == len(dest_file_list):
        infostr = "Copy of %d html support files succeeded" % len(src_file_list)
        print(infostr)
        LOGGER.info(infostr)
        # print '\n'.join(src_file_list)
    else:
        LOGGER.critical("FAILURE to cp html support files (%d source files,%d dest files)" % (
            len(src_file_list), len(dest_file_list)))
        sys.exit(1)

    # Now give read access to everyone, read-write to group, execute on directories
    try:
        group_id = grp.getgrnam(config_dict['chgrp_to']).gr_gid
    except KeyError:
        group_id = os.getegid()

    os.chmod(dest, 0o775)
    os.chown(dest, -1, group_id)

    for root, dirs, files in os.walk(dest):
        for dir in dirs:
            os.chmod(os.path.join(root, dir), 0o775)
            os.chown(os.path.join(root, dir), -1, group_id)
        for file in files:
            fname = os.path.join(root, file)
            os.chmod(fname, 0o664)
            os.chown(os.path.join(fname), -1, group_id)


# Return a dictionary of "generic Interactions" that can flow into the summary report
# Return an html table that can be placed at end of summary report
# Return html long report
# Return text that can be dumped to the log file
def gene_interaction_report(case_root, case, CheckInheritance):
    genepair_dict_file = os.path.join('.', 'casewide', 'DiGePred', '%s_all_gene_pairs_summary.json' % case)

    try:
        with open(genepair_dict_file, 'r') as fd:
            genepair_dict = json.load(fd)
    except FileNotFoundError:
        LOGGER.exception("Cannot open %s.  Did you run Souhrid's analysis scripts?" % genepair_dict_file)
        return None, None, None, None

    # genes_filename = os.path.join(case_root,"%s_genes.txt"%case)
    # try:
    #    with open(genes_filename) as fd:
    #       unstripped_genes = fd.readlines()
    # except:
    #    LOGGER.exception("Cannot open %s.  Did you run Souhrid's analysis scripts?"%genes_filename)
    #     return None,None,None,None

    geneInteractions = {}
    html_table = StringIO()
    html_report = StringIO()
    text_report = StringIO()

    structure_positive_genes = [gene.strip() for gene in list(sorted(genepair_dict.keys()))]

    pairs = {'direct': [], 'pathways': [], 'phenotypes': [], 'other': []}

    # We let the template html open the <table> with formatting as it prefers
    # print >>html_table, "<table>" 
    print("<thead><tr>", file=html_table)
    print("<th></th>", file=html_table)  # Top left is blank cell in header
    for gene1 in structure_positive_genes:
        print("<th>" + gene1 + "</th>", file=html_table)
    print("</tr></thead>", file=html_table)

    for gene1 in structure_positive_genes:
        print("%s:" % gene1, file=text_report)
        print('%s:' % gene1, file=html_report)
        print("<tr><td>%s</td>" % gene1, file=html_table)
        if gene1 in genepair_dict:
            print('<ul style="list-style: none;">', file=html_report)
        for gene2 in structure_positive_genes:
            print("<td>", file=html_table)
            if gene1 in genepair_dict and gene2 in genepair_dict[gene1] and (
                    (not CheckInheritance) or ('nherit' in genepair_dict[gene1][gene2])):
                pairs_hits = []
                if 'direct' in genepair_dict[gene1][gene2]:
                    pairs_hits.append('direct')
                if 'pwy' in genepair_dict[gene1][gene2]:
                    pairs_hits.append('pathways')
                if 'pheno' in genepair_dict[gene1][
                    gene2]:  # and genepair_dict[gene1][gene2]['Patient_Phenotype_Overlap'] > 0:
                    pairs_hits.append('phenotypes')
                # if genepair_dict[gene1][gene2]['Distance'] < 3 and genepair_dict[gene1][gene2]['Coexpression'] < 50:
                #    pairs_hits.append('other')
                if pairs_hits:
                    for pairs_hit in pairs_hits:
                        pairs[pairs_hit].append((gene1, gene2))
                        if pairs_hit != 'phenotypes':  # We don't get too excited for phenotypes at this point
                            if not gene1 in geneInteractions:
                                geneInteractions[gene1] = []
                            if not gene2 in geneInteractions[gene1]:
                                geneInteractions[gene1].append(gene2)
                    logging.debug("Gene Interactions: %s %s %s" % (gene1, gene2, str(pairs_hits)))
                    print(gene2, str(pairs_hits), file=text_report)
                    print("<li>", gene2 + ':', ", ".join(pairs_hits), "</li>", file=html_report)
                    print('<br>'.join(pairs_hits), file=html_table)
            print("</td>", file=html_table)
        print('</ul>', file=html_report)
        print("</tr>", file=html_table)
    # print("</table>", file=html_table)
    return geneInteractions, html_table.getvalue(), html_report.getvalue(), text_report.getvalue()




# =============================================================================
# Main logic here.  A period in the argument means the user wants to launch one mutation only,
# Whether one mutation, or the more typical set, we must have html/css/javascript statuc
# resources property installed in the destination tree
# =============================================================================

PDBMapSQLdb.set_access_dictionary(config_dict)

copy_html_css_javascript()

# CREATE the log/ webpage first, which includes details of everything done so far...
case_report_template_location = os.path.dirname(os.path.realpath(__file__))
jinja2_environment = Environment(loader=FileSystemLoader(case_report_template_location))
vustruct_logs_template = jinja2_environment.get_template("html/vustruct_logs_template.html")
# Variables with name case_ pertain to the creation of the final report summary outputs
case_jinja2 = {} # The dictionary of values that will be handed over to jinja2 to build out the case summary html
case_jinja2 = {
    'vustruct': vustruct.dict_for_jinja2(),
    'case': args.projectORstructure,
    'case_id': args.web_case_id,
    'case_uuid': args.strip_uuid,
    'date': time.strftime("%Y-%m-%d %H:%M") + ' USA Central Time (GMT - 6)',
}

# Change the displayed executable names from psb_ prefixes to vustruct_ prefixes
# Because we have renamed these command line components
# Also strip out any crazy leading path
for phase in ['preprocess', 'plan', 'launch', 'monitor', 'report']:
    if phase in case_jinja2['vustruct']:
        python_exe_name = os.path.basename(case_jinja2['vustruct'][phase]['executable'])
        vustruct_display_name = python_exe_name.replace("psb_","vustruct_")
        case_jinja2['vustruct'][phase]['executable'] = vustruct_display_name
# We need to get rid of log/ prefixes, because we generate index.html
# down in the log directory
for command_line_module in case_jinja2['vustruct'].keys():
    if 'log_filename' in case_jinja2['vustruct'][command_line_module]:
        log_filename = case_jinja2['vustruct'][command_line_module]['log_filename']
        if log_filename: # Then add log_basename to in-memory dictionary
            website_filelist.append(log_filename)
            case_jinja2['vustruct'][command_line_module]['log_basename'] = \
                os.path.basename(log_filename)
    if 'input_filename' in case_jinja2['vustruct'][command_line_module]:
        input_filename = case_jinja2['vustruct'][command_line_module]['input_filename']
        if input_filename: # Then add input_basename to in-memory dictionary
            website_filelist.append(input_filename)
            case_jinja2['vustruct'][command_line_module]['input_basename'] = \
                os.path.basename(input_filename)


# The case_summary worksheet is an xlsx file that _drafts_ a report
# for the clinical team.  It is a timesaver, and not final.
# The idea is that the structural biologist will use this file as
# a starting point for the final report
case_workbook = Workbook()
case_summary_worksheet = case_workbook.active
case_summary_worksheet.title = args.web_case_id if args.web_case_id else args.projectORstructure
case_summary_worksheet.name = "Summary"
case_summary_headers = [
"Gene","Uniprot ID","Variant", "ddG", "PathProx 3D hotspot vs Clinvar", "PathProx 3D hotspot vs COSMIC",
"COSMIS", "ScanNet PPI", "Alpha Missense", "Musite PTM", "Predicted Digenic Interactions","Worth pursuing?","side notes"]
case_summary_worksheet.append(case_summary_headers)
case_summary_worksheet_widths = [10,8,10,15,15,10,10,10,10,10,10,10,8]

ddG_cartesian_worksheet = case_workbook.create_sheet("ddG Cartesian")
ddG_monomer_worksheet = case_workbook.create_sheet("ddG Monomer")

# We collect ddGs and for each gene.transcript,variant, we map to a dictionary of structure IDs: ddG calculation
# Typehints - a work in progress
# ddG_monomer_xref = dict[tuple[string,string,string],dict[str,float]]
# ddG_cartesian_xref = dict[tuple[string,string,string],dict[str,float]]
ddG_monomer_xref = {}
ddG_cartesian_xref = {}
def ddg_xref_add(variant_row, variant_isoform_details) -> None:
    """
    Build out a cross reference between each variant and the structures which have ddGs
    This is used to later create ddG Cartesian and Monomer .xlsx worksheets in summary
    """
    for ddG_monomer_or_cartesian, ddG_xref in {
        'ddG Monomer': ddG_monomer_xref,
        'ddG Cartesian': ddG_cartesian_xref,
        }.items():
    
        variant_key = (variant_row['gene'],
                       variant_row['unp'],
                       variant_row['mutation'])

        ddG_xref[variant_key] = {}
        for struct_tuple, ddg_value in variant_isoform_details[ddG_monomer_or_cartesian].items():
            # import pdb; pdb.set_trace()
            _method, _structure_id, _chain, _mers = struct_tuple

            ddG_xref[(variant_row['gene'],
                              variant_row['unp'],
                              variant_row['mutation'])][_structure_id + "." + _chain] = ddg_value

def ddg_populate_worksheet(ddg_worksheet: openpyxl.worksheet.worksheet.Worksheet, ddG_xref: Dict):
    """
    Create a single worksheet with variants down the first column
    and structures along the top row
    """
    ddg_worksheet_columns = {}

    for variant, ddg_results in ddG_xref.items():
        for structure_id, ddg_result in ddg_results.items():
            # print("%s %s -> %s" % (variant, model_id, ddg_result))
            if structure_id not in ddg_worksheet_columns:
                ddg_worksheet_columns[structure_id] = 4 + len(ddg_worksheet_columns)
    
    # for structure_id, column in ddg_worksheet_columns.items():
    #    print ("%02d %s" % (column, structure_id))
    
    ddg_worksheet.cell(row=1, column=1).value='Gene'
    ddg_worksheet.cell(row=1, column=2).value='Uniprot'
    ddg_worksheet.cell(row=1, column=3).value='Variant'
    for structure_id, column in  ddg_worksheet_columns.items():
        cell = ddg_worksheet.cell(row=1, column=column)
        cell.value= structure_id
        # print("%s %s" % (column, cell.alignment))
        # cell.value = headers[key]
        cell.alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=90)
    
    ddg_worksheet.row_dimensions[1].height = 180
    
    current_row = 2
    for variant, ddg_results in ddG_xref.items():
        ddg_worksheet.cell(row=current_row,column=1).value = variant[0]
        ddg_worksheet.cell(row=current_row,column=2).value = variant[1]
        ddg_worksheet.cell(row=current_row,column=3).value = variant[2]
        for structure_id, ddg_result in ddg_results.items():
            ddg_worksheet.cell(row=current_row,column=ddg_worksheet_columns[structure_id]).value = ddg_result
        current_row += 1





assert len(case_summary_worksheet_widths) == len(case_summary_headers)

for col in range(1,1+len(case_summary_headers)):
    case_summary_worksheet.cell(1,col).font = openpyxl.styles.Font(name = "Arial", bold=True)
    case_summary_worksheet.cell(1,col).alignment = Alignment(wrap_text=True)

dim_holder = DimensionHolder(worksheet=case_summary_worksheet)

for col in range(case_summary_worksheet.min_column, case_summary_worksheet.max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(case_summary_worksheet, min=col, max=col, width=case_summary_worksheet_widths[col-1])

# dim_holder[1] = RowDimension(case_summary_worksheet, min=1, max=1, height=10)
case_summary_worksheet.column_dimensions = dim_holder

case_summary_worksheet.row_dimensions[1].height = 30
# case_workbook.save('/tmp/crap.xlsx')


class variant_isoform_helper:
    """
    This set of functions helps tease apart components of the isoform summaries returned by rep_one_variant
    We can encapsulate all of this better in future, with a new variant_isoform_summary class...

    """
    @staticmethod
    def min_max_cosmis(variant_isoform_summaries: list[dict]) -> tuple[float, float]:
        min_cosmis_score = min(
            [variant_isoform_summary['cosmis_score'] for variant_isoform_summary in
             variant_isoform_summaries \
             if 'cosmis_score' in variant_isoform_summary and variant_isoform_summary[
                 'cosmis_score'] is not None],
            default=None)

        max_cosmis_score = max(
            [variant_isoform_summary['cosmis_score'] for variant_isoform_summary in
             variant_isoform_summaries \
             if 'cosmis_score' in variant_isoform_summary and variant_isoform_summary[
                 'cosmis_score'] is not None],
            default=None)

        return min_cosmis_score, max_cosmis_score

    @staticmethod
    def min_max_alphamissense(variant_isoform_summaries: list[dict]) -> tuple[float, float]:
        min_alphamissense_score = min(
            [variant_isoform_summary['alphamissense_score'] for variant_isoform_summary in
             variant_isoform_summaries \
             if 'alphamissense_score' in variant_isoform_summary and variant_isoform_summary[
                 'alphamissense_score'] is not None],
            default=None)

        max_alphamissense_score = max(
            [variant_isoform_summary['alphamissense_score'] for variant_isoform_summary in
             variant_isoform_summaries \
             if 'alphamissense_score' in variant_isoform_summary and variant_isoform_summary[
                 'alphamissense_score'] is not None],
            default=None)

        return min_alphamissense_score, max_alphamissense_score

    @staticmethod
    def min_max_ScanNetPPI(variant_isoform_summaries: list[dict]) -> tuple[float, float]:
        max_ScanNetPPI = max(
            [variant_isoform_summary['ScanNetPPI'][0]['Binding site probability'] for variant_isoform_summary in
             variant_isoform_summaries \
             if 'ScanNetPPI' in variant_isoform_summary and 
                 variant_isoform_summary['ScanNetPPI'] is not None and 
                 len(variant_isoform_summary['ScanNetPPI']) > 0 ],
            default=None)

        min_ScanNetPPI = min(
            [variant_isoform_summary['ScanNetPPI'][0]['Binding site probability'] for variant_isoform_summary in
             variant_isoform_summaries \
             if 'ScanNetPPI' in variant_isoform_summary and 
                 variant_isoform_summary['ScanNetPPI'] is not None and 
                 len(variant_isoform_summary['ScanNetPPI']) > 0 ],
            default=None)

        return min_ScanNetPPI, max_ScanNetPPI
    
    @staticmethod
    def min_max_MusiteDeepPTM(variant_isoform_summaries: list[dict]) -> tuple[float, float]:
        # import pdb; pdb.set_trace()
        _musiteDeep_probabilities = []
        for variant_isoform_summary in variant_isoform_summaries:
            if 'MusiteDeepPTM' in variant_isoform_summary:
                for ptm_record in variant_isoform_summary['MusiteDeepPTM']:
                    _musiteDeep_probabilities.append(ptm_record['PTM Probability'])
        max_MusiteDeepPTM = max(_musiteDeep_probabilities, default=None)
        min_MusiteDeepPTM = min(_musiteDeep_probabilities, default=None)

        return min_MusiteDeepPTM, max_MusiteDeepPTM

def write_case_report_html(case_jinja2_dictionary) -> None:
    """ The top level "case report" is the output index.html
        For the entire run.  Depending on progress of the pipeline
        case_jinja2_dictionary includes preprocess log details, and then vustruct gene lins
        Then outlinks to individual variant reports
        param case_jinja2_dictionary - all values for the jinja2 processor
    """
    case_report_template_location = os.path.dirname(os.path.realpath(__file__))
    jinja2_environment = Environment(loader=FileSystemLoader(case_report_template_location))
    def basename(path):
        return os.path.basename(path)
    jinja2_environment.filters['filter_basename'] = basename
    case_report_template = jinja2_environment.get_template("case_report_template.html")
    html_out = case_report_template.render(case_jinja2_dictionary)
    # args.projectORstructure is an entire project UDN124356

    case_summary_filename = os.path.join(case_root_dir,
                                         "%s.html" % args.projectORstructure)
    website_filelist.append(case_summary_filename)

    with open(case_summary_filename, "w") as f:
        f.write(html_out)
    return case_summary_filename

def write_case_website_tar_zip(case_summary_filename):
    if website_filelist:
        try:
            os.remove('index.html')  # First time through will gen an exception.  That's A-OK
        except OSError:  # A-OK if file is not arleady there
            pass
        try:
            index_html_symlink = case_summary_filename
            # If our destiny is to later rename the main .html filename with the tar --transform to come, then
            # we need to repoing index.html to the forthcoming renamed file
            if args.strip_uuid and args.web_case_id:
                index_html_symlink = args.web_case_id + '.html'
            # The symlink fails when we are running in some vm environments - so we just copy in those cases
            os.symlink(index_html_symlink, 'index.html')
        except PermissionError:
            LOGGER.info("Creating index.html as symlink failed in VM.  Attempting simpler cp operations")
            shutil.copy(src=case_summary_filename, dst='index.html')
            pass
        website_filelist.append('index.html')

        website_filelist_filename = os.path.join(case_root_dir,
                                                 "%s_website_files.list" % args.projectORstructure)  # The argument is an entire project UDN124356
        with open(website_filelist_filename, "w") as f:
            f.write('\n'.join((os.path.relpath(
                website_file,  # os.path.realpath(website_file),
                (os.path.realpath(os.path.join(config_dict['output_rootdir'], config_dict['collaboration']))))
                for website_file in website_filelist)))
        LOGGER.info("A filelist for creating a website is in %s", website_filelist_filename)
        website_zip_filename = "%s.zip" % args.projectORstructure
        if args.tar_only:
            LOGGER.info("--tar_only requested.  %s will not be created." , website_zip_filename);
        else:
            pkzip_maker = 'rm -f %s; cd ..; cat %s | zip -r@ %s > %s.stdout; cd -' % (
                website_zip_filename,
                os.path.join(args.projectORstructure, website_filelist_filename),
                os.path.join(args.projectORstructure, website_zip_filename),
                os.path.join(args.projectORstructure, website_zip_filename))
            LOGGER.info("Creating .zip website file with: %s", pkzip_maker)
            subprocess.call(pkzip_maker, shell=True)

        # The temp filename has a hideous UTC timestamp on it
        website_tar_temp_filename = "%s_%s.tar.gz" % (
            args.projectORstructure, 
            datetime.datetime.now().replace(microsecond=0).isoformat().translate(str.maketrans('','',string.punctuation)))

        website_tar_filename = "%s.tar.gz" % args.projectORstructure
        tar_transformer = ''
        if args.strip_uuid:
            # Without some post processing, Some filenames stored in the website will contain the full uuid, where we only want the
            # caseID in the file name
            # We "want" this in the end:
            #    external_user_EMAILTEST_18f99df6-6880-44b0-86c6-da5c606e8c18/EMAILTEST.html
            # but the original component filename must be transformed from this:
            #    external_user_EMAILTEST_18f99df6-6880-44b0-86c6-da5c606e8c18/external_user_EMAILTEST_18f99df6-6880-44b0-86c6-da5c606e8c1.html
            case_id = args.web_case_id
            uuid = args.strip_uuid
            search_string = f"/external_user_{case_id}_{args.strip_uuid}"
            replace_string = f"/{case_id}"
            tar_transformer = f"--transform 's[{search_string}[{replace_string}[g' --show-transformed-names"
        tar_maker = 'cd ..; tar cvzf %s --files-from %s %s --mode=\'a+rX,go-w,u+w\' > %s.stdout; cd -; mv %s %s; mv %s.stdout %s.stdout' % (
            os.path.join(args.projectORstructure, website_tar_temp_filename),
            os.path.join(args.projectORstructure, website_filelist_filename),
            tar_transformer,
            os.path.join(args.projectORstructure, website_tar_temp_filename),
            website_tar_temp_filename,website_tar_filename,
            website_tar_temp_filename,website_tar_filename)
        LOGGER.info("Creating .tar website file with: %s", tar_maker)
        subprocess.call(tar_maker, shell=True)

        final_message = "Compressed website files are in "
        if args.tar_only:
            final_message += website_tar_filename
        else:
            final_message += "%s and %s" % ( website_zip_filename, website_tar_filename)

        LOGGER.info(final_message);



    # Create and save the supplemental written files to aid presentation
    # - summary of summaries case spreadsheet

# directly from a single mutation output file of psb_plan.py
if oneMutationOnly:
    if args.slurm:
        LOGGER.warning("--slurm option is ignored when only one report is requested")
    template_vars = report_one_variant_one_isoform(args.projectORstructure, args.workstatus)
    if template_vars:  # The usual case, we had some Pathprox outputs
        print("Single mutation report saved to %s/.pdf/.wkhtml.pdf" % template_vars['html_fname'])
    else:
        print("Due to lack of pathprox outputs, no html (or pdf) reports were created from %s" % args.workstatus)
else:
    _plan_failed = True
    _launch_failed = True

    # status_header conveys information about the pipeline and could be positive
    # or negative 
    case_jinja2['status_header'] = None
    case_jinja2['module_failure_info'] = None 


    df_all_mutations = pd.DataFrame()
    vustruct_csv_data = ""
    case_vustruct_filename = "Not Created"
    module_failure_info = None # Initialize only if a preprocess or plan phase failed

    def textlist_to_html(text_list: List[str]) -> str:
        html_codeblock = ""
        for textline in text_list:
            html_codeblock += html.escape(textline) + "\n"
        return html_codeblock

    def basename_psb_vustruct(filename: str):
        return os.path.basename(filename).replace("psb","vustruct")

    if vustruct.preprocess_failed: # Then we can go no farther
         case_jinja2['status_header'] = "Preprocessor %s failed" % (
             basename_psb_vustruct(vustruct.preprocess['executable']),
             )
         case_jinja2['module_failure_info'] = textlist_to_html(vustruct.preprocess_failure_info)
         case_jinja2['preprocess_failed'] = True
         LOGGER.info(case_jinja2['status_header'])
         _case_summary_filename = write_case_report_html(case_jinja2)
         write_case_website_tar_zip(_case_summary_filename)
         sys.exit(0) # It's A-OK that psb_rep is reporting on a failed case start

    if vustruct.plan_failed: # Then we can go no farther
         case_jinja2['status_header'] = "Plan module %s failed" % (
             basename_psb_vustruct(vustruct.plan['executable']),
             )
         case_jinja2['module_failure_info'] = textlist_to_html(vustruct.plan_failure_info)
         case_jinja2['plan_failed'] = True
         _case_summary_filename = write_case_report_html(case_jinja2)
         write_case_website_tar_zip(_case_summary_filename)
         sys.exit(0) # It's A-OK that psb_rep is reporting on a failed case start

    # We now press onwards showing the entire vustruct file as well and, when available, outlinkds to individual reports

    case_vustruct_filename = os.path.join(case_root_dir,
                                "%s_vustruct.csv" % args.projectORstructure)  # The argument is an entire project UDN124356
    msg = "Retrieving project mutations from %s" % case_vustruct_filename
    if not infoLogging:
        print(msg)
    LOGGER.info(msg)

    with open(case_vustruct_filename,'r') as f:
        vustruct_csv_data = f.read()
    df_all_mutations = pd.read_csv(StringIO(vustruct_csv_data), sep=',', index_col=None, keep_default_na=False, encoding='utf8',
                                   comment='#', skipinitialspace=True)
    df_all_mutations.fillna('NA', inplace=True)
    website_filelist.append(os.path.basename(case_vustruct_filename))

    vustruct_logs_info = {
        'case_vustruct_filename': os.path.basename(case_vustruct_filename),
        'case_vustruct_csv_data': vustruct_csv_data,
        'case_vustruct_df': df_all_mutations,
        'refreshFlag': args.embed_refresh,
        'vustruct': case_jinja2['vustruct']
    }

    if args.slurm:
        slurm_directory = os.path.join(case_root_dir, "slurm")
        slurm_file = os.path.join(slurm_directory, "psb_reps.slurm")
        slurm_stdout_directory = os.path.join(slurm_directory, "stdout")
        if not os.path.exists(slurm_directory):  # make the containing directory if needed
            os.makedirs(slurm_directory)
        if not os.path.exists(slurm_stdout_directory):  # make the stdout directory if needed
            os.makedirs(slurm_stdout_directory)
        jobCount = len(df_all_mutations)
        msg = "Slurm script to run %d reports for %s is %s" % (jobCount, args.projectORstructure, slurm_file)
    else:  # not a slurm run - so just do one report after another- normal case
        msg = "Reporting on all %d case=%s mutations" % (len(df_all_mutations), args.projectORstructure)
    if not infoLogging:
        print(msg)
    LOGGER.info(msg)

    slurm_array = []

    # Determine the formatting style of the main case page.
    # If CHROMs and POSs are availabine in the source data, THEN we create a heirarchical grouping above the
    # individual transcript processing.
    genome_variants = defaultdict(list)
    # mutation_summaries = [] # Used to populate old format report, one line per gene
    genome_headers = []  # Used for new format report, with one line per gene, expandable to transcript isoforms

    def _format_min_max(minval, maxval, decimal_places=2, is_pctg=False) -> str:
        if minval is None or math.isnan(minval): # Then they both have to be nans
            return ""

        _fmtstr = "%0." + str(decimal_places) + "f"

        if is_pctg:
          minval *= 100.0
          maxval *= 100.0
          _fmtstr += "%%"
            
        # Don't output a "-0.00"
        
        minval = round(minval, decimal_places)
        if minval == 0.0:
            minval = 0.0
        maxval = round(maxval, decimal_places)
        if maxval == 0.0:
            maxval = 0.0


        try:
            # If they both round to the same string, then 
            # of course just output one, not a range
            if (_fmtstr % minval) == (_fmtstr % maxval):
                return _fmtstr % maxval
            else: # We have a range
                return (_fmtstr + " - " + _fmtstr) % (minval, maxval)
        except:
            return ""
    

    def fetch_gene_id(uniprot_id: str):

        query = """\
SELECT ID FROM Idmapping 
where Id_Type = 'GeneID' and unp = %(unp)s"""

        with PDBMapSQLdb(config_dict) as sql_connection:
            unp_base = uniprot_id.split('-')[0]
            sql_connection.execute(query, {'unp': unp_base})
            row = sql_connection.fetchone()
            if not row or len(row) != 1:
                LOGGER.warning("unp %s is invalid, or lacks a GeneID entry in the idmapping file" % unp_base)
                return ""
            else:
                return row[0].strip()


    # When we have chrom+pos+change columns in the input, then make collapsible rows
    # for each entry, where the transcripts are shown when the row is expanded.
    if {'chrom', 'pos', 'change'}.issubset(df_all_mutations.columns) and not args.slurm:
        for index, row in df_all_mutations.iterrows():
            chrom_pos_change = (row['chrom'], row['pos'], row['change'])
            genome_variants[chrom_pos_change].append((index, row))

        for chrom_pos_change in genome_variants:
            # Crude - but get at the original missense row
            # under which our protein transcript isoform are grouped
            first_row = genome_variants[chrom_pos_change][0][1]
            msg = "Reporting on: %s %s" % (first_row['gene'], chrom_pos_change)
            if not infoLogging:
                print(msg)
            LOGGER.info(msg)
            genome_header = {
                'Gene': first_row['gene'],
                'Chrom': first_row['chrom'],
                'Pos': first_row['pos'],
                'Change': first_row['change'],
                'disease1_pp Min': None,
                'disease1_pp Max': None,
                'disease2_pp Min': None,
                'disease2_pp Max': None,
                'ddG Monomer Min': None,
                'ddG Monomer Max': None,
                'ddG Cartesian Min': None,
                'ddG Cartesian Max': None,
                'alphamissense_score': None,
                'ScanNetPPI': None,
                'MusiteDeepPTM': None,
                'variant_isoform_summaries': {}
            }

            variant_isoform_summaries = []
            genome_header['variant_isoform_summaries'] = variant_isoform_summaries

            for index, genome_variant_row in genome_variants[chrom_pos_change]:
                msg = "%s %-10s %-10s %-6s" % (
                    "      ", genome_variant_row['refseq'], genome_variant_row['mutation'], genome_variant_row['unp'])
                if not infoLogging:
                    print(msg)
                LOGGER.info(msg)
                variant_directory = "%s_%s_%s" % (
                    genome_variant_row['gene'], genome_variant_row['refseq'], genome_variant_row['mutation'])
                variant_isoform_summary, variant_isoform_details, additional_website_files = report_one_variant_one_isoform(
                    args.projectORstructure,
                    case_root_dir= case_root_dir,
                    variant_directory_segment= variant_directory,
                    parent_report_row=genome_variant_row,
                    local_log_level= "debug" if args.debug else "info" if args.verbose else "warn",
                    vustruct_pipeline_launched= bool(case_jinja2['vustruct']['launch']['executable']),
                    config_pathprox_dict=config_pathprox_dict
                )
                if variant_isoform_summary:
                    ddg_xref_add(genome_variant_row, variant_isoform_details)
                    website_filelist.extend(additional_website_files)
                    variant_isoform_summary['#'] = index
                    if not variant_isoform_summary['Error']:
                        msg = "%s %s %s report saved to %s" % (
                            genome_variant_row['gene'], genome_variant_row['refseq'], genome_variant_row['mutation'],
                            variant_isoform_summary['html_filename'])
                        if not infoLogging:
                            print(msg)
                        LOGGER.info(msg)
                    variant_isoform_summary['Gene'] = genome_variant_row['gene']

                    variant_isoform_summary['Unp'] = genome_variant_row['unp']
                    variant_isoform_summary['gene_id'] = fetch_gene_id(variant_isoform_summary['Unp'])
                    transcript_list = genome_variant_row['transcript'].split(';')
                    variant_isoform_summary['Transcript'] = transcript_list[0]
                    if len(transcript_list) > 1:
                        variant_isoform_summary['Transcript'] += '...'
                    variant_isoform_summary['Refseq'] = genome_variant_row['refseq']
                    variant_isoform_summary['Mutation'] = genome_variant_row['mutation']
                    variant_isoform_summaries.append(variant_isoform_summary)




                else:
                    msg = "Due to lack of pathprox or ddG outputs, %s %s %s has no html (or pdf) report" % (
                        genome_variant_row['gene'], genome_variant_row['refseq'], genome_variant_row['mutation'])
                    if not infoLogging:
                        print(msg)
                    LOGGER.info(msg)

            if variant_isoform_summaries:

                # The Gene ID will be same for all isoform summaries
                genome_header['gene_id'] = variant_isoform_summaries[0]['gene_id']

                genome_header['ddG Monomer Min'] = min(
                    [variant_isoform_summary['ddG Monomer Min'] for variant_isoform_summary in
                     variant_isoform_summaries \
                     if 'ddG Monomer Min' in variant_isoform_summary and variant_isoform_summary[
                         'ddG Monomer Min'] is not None],
                    default=None)
                genome_header['ddG Monomer Max'] = max(
                    [variant_isoform_summary['ddG Monomer Max'] for variant_isoform_summary in
                     variant_isoform_summaries \
                     if 'ddG Monomer Max' in variant_isoform_summary and variant_isoform_summary[
                         'ddG Monomer Max'] is not None],
                    default=None)

                genome_header['ddG Cartesian Min'] = min(
                    [variant_isoform_summary['ddG Cartesian Min'] for variant_isoform_summary in
                     variant_isoform_summaries \
                     if 'ddG Cartesian Min' in variant_isoform_summary and variant_isoform_summary[
                         'ddG Cartesian Min'] is not None],
                    default=None)
                genome_header['ddG Cartesian Max'] = max(
                    [variant_isoform_summary['ddG Cartesian Max'] for variant_isoform_summary in
                     variant_isoform_summaries \
                     if 'ddG Cartesian Max' in variant_isoform_summary and variant_isoform_summary[
                         'ddG Cartesian Max'] is not None],
                    default=None)

                genome_header['AA_Len Min'] = min(
                    [variant_isoform_summary['AA_len'] for variant_isoform_summary in variant_isoform_summaries \
                     if 'AA_len' in variant_isoform_summary and variant_isoform_summary['AA_len'] is not None],
                    default=None)

                genome_header['AA_Len Max'] = max(
                    [variant_isoform_summary['AA_len'] for variant_isoform_summary in variant_isoform_summaries \
                     if 'AA_len' in variant_isoform_summary and variant_isoform_summary['AA_len'] is not None],
                    default=None)

                genome_header['disease1_pp Min'] = min(
                    [variant_isoform_summary['disease1_pp Min'] for variant_isoform_summary in
                     variant_isoform_summaries \
                     if 'disease1_pp Min' in variant_isoform_summary and variant_isoform_summary[
                         'disease1_pp Min'] is not None],
                    default=None)
                genome_header['disease1_pp Max'] = max(
                    [variant_isoform_summary['disease1_pp Max'] for variant_isoform_summary in
                     variant_isoform_summaries \
                     if 'disease1_pp Max' in variant_isoform_summary and variant_isoform_summary[
                         'disease1_pp Max'] is not None],
                    default=None)

                genome_header['disease2_pp Min'] = min(
                    [variant_isoform_summary['disease2_pp Min'] for variant_isoform_summary in
                     variant_isoform_summaries \
                     if 'disease2_pp Min' in variant_isoform_summary and variant_isoform_summary[
                         'disease2_pp Min'] is not None],
                    default=None)
                genome_header['disease2_pp Max'] = max(
                    [variant_isoform_summary['disease2_pp Max'] for variant_isoform_summary in
                     variant_isoform_summaries \
                     if 'disease2_pp Max' in variant_isoform_summary and variant_isoform_summary[
                         'disease2_pp Max'] is not None],
                    default=None)

                _min_cosmis_score, _max_cosmis_score = \
                    variant_isoform_helper.min_max_cosmis(variant_isoform_summaries)
                genome_header['cosmis_score'] = \
                    _format_min_max(_min_cosmis_score, _max_cosmis_score, 1)

                _min_alphamissense_score, _max_alphamissense_score = \
                    variant_isoform_helper.min_max_alphamissense(variant_isoform_summaries)

                # if _min_alphamissense_score == _max_alphamissense_score:
                #     genome_header['alphamissense_score'] = "%0.2f" % _max_alphamissense_score
                # else:
                #     genome_header['alphamissense_score'] = "%0.2f-%0.2f" % (
                #         _min_alphamissense_score,_max_alphamissense_score)
                genome_header['alphamissense_score'] = \
                    _format_min_max(_min_alphamissense_score,_max_alphamissense_score,2)
              

                _min_ScanNetPPI, _max_ScanNetPPI = \
                    variant_isoform_helper.min_max_ScanNetPPI(variant_isoform_summaries)

                if _max_ScanNetPPI is None:
                    genome_header['ScanNetPPI'] = None
                else:
                    genome_header['ScanNetPPI'] = _format_min_max(_min_ScanNetPPI,_max_ScanNetPPI,0,True)
    
                genome_header['MusiteDeepPTM'] = 0.0

                _min_MusiteDeepPTM,_max_MusiteDeepPTM = \
                    variant_isoform_helper.min_max_MusiteDeepPTM(variant_isoform_summaries)
                
                if _max_MusiteDeepPTM is None:
                    genome_header['MusiteDeepPTM'] = None
                else:
                    genome_header['MusiteDeepPTM'] = _format_min_max(_min_MusiteDeepPTM,_max_MusiteDeepPTM,0,True)

                genome_header['Error'] = max(
                    [variant_isoform_summary['Error'] for variant_isoform_summary in variant_isoform_summaries \
                     if 'Error' in variant_isoform_summary and variant_isoform_summary['Error'] is not None],
                    default=None)

                genome_headers.append(genome_header)

    else:  # There are no chrom positions in this case.  Use the old format
        variant_isoform_summaries = []  # One line per variant
        for index, variant_row in df_all_mutations.iterrows():
            msg = "%s %-10s %-10s %-6s" % (
                "Generating slurm entry for" if args.slurm else "Reporting on", variant_row['gene'],
                variant_row['refseq'], variant_row['mutation'])
            if not infoLogging:
                print(msg)
            LOGGER.info(msg)

            if 'RefSeqNotFound_UsingGeneOnly' in variant_row['refseq']:
                variant_row['refseq'] = 'NA'
            # gene_refseq_mutation,structure_report_filename,workstatus_filename,dropped_structure_filenames = paths_for_variant_row(row)

            if args.slurm:
                slurm_array.append("./psb_rep.py -c %s -u %s %s %s" % (
                    args.config, args.userconfig, structure_report_filename, workstatus_filename))
            else:
                variant_directory = "%s_%s_%s" % (variant_row['gene'], variant_row['refseq'], variant_row['mutation'])
                variant_isoform_summary, variant_isoform_details,additional_website_files = report_one_variant_one_isoform(
                    project=args.projectORstructure,
                    case_root_dir=case_root_dir,
                    variant_directory_segment=variant_directory, 
                    parent_report_row=variant_row,
                    local_log_level= "debug" if args.debug else "info" if args.verbose else "warn",
                    vustruct_pipeline_launched= bool(case_jinja2['vustruct']['launch']['executable']),
                    config_pathprox_dict=config_pathprox_dict
                    )

                try:
                    case_workbook.create_sheet(variant_directory)
                except ValueError as e:
                    cleaned_variant_directory_name = variant_directory.encode('utf-8','ignore').decode("utf-8")
                    case_workbook.create_sheet(cleaned_variant_directory_name.translate(str.maketrans('','',string.punctuation)))

                if variant_isoform_summary:
                    ddg_xref_add(variant_row, variant_isoform_details)
                    website_filelist.extend(additional_website_files)
                    variant_isoform_summary['#'] = index
                    if not variant_isoform_summary['Error']:
                        msg = "%s %s %s report saved to %s" % (
                            variant_row['gene'], variant_row['refseq'], variant_row['mutation'],
                            variant_isoform_summary['html_filename'])
                        if not infoLogging:
                            print(msg)
                        LOGGER.info(msg)
                    variant_isoform_summary['Gene'] = variant_row['gene']
                    variant_isoform_summary['Unp'] = variant_row['unp']
                    variant_isoform_summary['gene_id'] = fetch_gene_id(variant_isoform_summary['Unp'])
                    variant_isoform_summary['Refseq'] = variant_row['refseq']
                    variant_isoform_summary['Mutation'] = variant_row['mutation']
                    variant_isoform_summaries.append(variant_isoform_summary)
                else:
                    msg = "Due to lack of pathprox or ddG outputs, %s %s %s has no html (or pdf) report" % (
                        variant_row['gene'], variant_row['refseq'], variant_row['mutation'])
                    if not infoLogging:
                        print(msg)
                    LOGGER.info(msg)

        if args.slurm:
            with open(slurm_file, "w") as slurmf:
                slurmf.write("""\
    #!/bin/sh
    #
    # Project        : PSB Pipeline
    # Filename       : %s
    # Generated By   : %s
    # For case       : %s
    # Organization   : Vanderbilt Genetics Institute,
    #                : Program in Personalized Structural Biology,
    #                : Department of Biomedical Informatics,
    #                : Vanderbilt University
    # Generated on   : %s
    # Description    : Runs a python script on a cluster to launch psb_rep.py for 
    #                : each and all mutations in parallel
    #===============================================================================
    # Slurm Parameters
    """ % (slurm_file, __file__, args.projectORstructure, str(datetime.datetime.now())))

                slurmDict = slurmParametersReport
                slurmDict['output'] = os.path.join(slurm_stdout_directory,
                                                   "%s_%%A_%%a_psb_rep.out" % args.projectORstructure)
                slurmDict['job-name'] = "%s_psb_reps" % args.projectORstructure

                jobCount = len(slurm_array)
                if jobCount > 1:
                    slurmDict['array'] = "0-%d" % (jobCount - 1)

                # We can set a max on the number of simultaneously running jobs in the array.  For now, don't do this
                # if jobCount > 10:
                #   slurmDict['array'] += "%%10"

                for slurm_field in ['job-name', 'mail-user', 'mail-type', 'ntasks', 'time', 'mem', 'account', 'output',
                                    'array']:
                    if slurm_field in slurmDict:
                        slurmf.write("#SBATCH --%s=%s\n" % (slurm_field, slurmDict[slurm_field]))

                if jobCount > 1:
                    slurmf.write("""
    echo "SLURM_ARRAY_TASKID="$SLURM_ARRAY_TASKID
    """)
                slurmf.write("""
    echo "SLURM_JOBID="$SLURM_JOBID
    echo "SLURM_JOB_NODELIST"=$SLURM_JOB_NODELIST
    echo "SLURM_NNODES"=$SLURM_NNODES
    # echo "SLURMTMPDIR="$SLURMTMPDIR
    echo "SLURM_SUBMIT_DIR = "$SLURM_SUBMIT_DIR
    
    """)

                cwd = os.getcwd()

                slurmf.write("""
    cd %s
    if [ $? != 0 ]; then
    echo Failure at script launch: Unable to change to directory %s
    exit 1
    fi
    """ % (cwd, cwd))

                if jobCount == 1:
                    # No need to fiddle with the slurm case statement
                    slurmf.write(slurm_array[0])
                else:
                    slurmf.write("\ncase $SLURM_ARRAY_TASK_ID in\n")
                    slurm_array_id = 0
                    for launchstring in slurm_array:
                        # Write out the "case" tag
                        slurmf.write("%d)\n" % slurm_array_id)
                        # Save this to put back later
                        slurm_array_id += 1
                        slurmf.write("echo Command: '%s'\n" % launchstring)
                        slurmf.write(launchstring)
                        # end the case tag
                        slurmf.write("\n;;\n\n")
                    slurmf.write("esac\n")
            msg = "Created slurm script to launch all psb_rep.py processes for this case: %s" % slurm_file
            if args.verbose:
                LOGGER.info(msg)
            else:
                print(msg)


    # It can easily be the case that we arrive here WITHOUT data because the pipeline is not launched.
    # In that case, we need to give an update on pipeline progress and NOT a table of results.
    # In this case, the following "if" will be false
    early_or_fail_message = ""
    if not (genome_headers or variant_isoform_summaries):  # Excellent, we have a home page report for many variants
        if not vustruct.plan['executable']:
            early_or_fail_message = "The plan phase has not been run.  Check logs for errors"

    ##########################################################################
    # Was there a generated .html file from DiGePred?
    ##########################################################################
    digepred_html_filename = os.path.join('casewide', 'DiGePred',
                                          '%s_all_gene_pairs_summary.html' % (
                                              args.projectORstructure,))

    if os.path.exists(os.path.join(case_root_dir, digepred_html_filename)):
        website_filelist.append(digepred_html_filename)
        LOGGER.info("Integrating DiGePred html: %s", digepred_html_filename)
    else:
        LOGGER.warning("DiGePred %s missing.  Did you run DiGePred?", digepred_html_filename)
        digepred_html_filename = None

    digepred_gene_pairs = []
    if digepred_html_filename:
        digepred_csv_filename = os.path.join('casewide', 'DiGePred',
                                             '%s_all_gene_pairs_digenic_metrics.csv' % (
                                                 args.projectORstructure,))

        digepred_metrics_df = pd.DataFrame()
        if os.path.exists(os.path.join(case_root_dir, digepred_csv_filename)):
            LOGGER.info("Integrating DiGePred csv: %s", digepred_csv_filename)
            digepred_metrics_df = pd.read_csv(digepred_csv_filename, sep=',')
            LOGGER.info("%d rows read", len(digepred_metrics_df))
            digepred_metrics_df.sort_values(by=['digenic score'], ascending=False, inplace=True, ignore_index=True)
        else:
            LOGGER.critical("DiGePred %s missing.  Alarming, as .html file was found", digepred_csv_filename)
            del digepred_csv_filename

        # Now prepare the entries which will go to the html table
        # For now, max 20 or all
        max_rows = min(20, len(digepred_metrics_df))
        for index, row in digepred_metrics_df.head(max_rows).iterrows():
            gene_pair_dict = {}
            for key in ['gene A', 'gene B', 'digenic score']:
                gene_pair_dict[key] = row[key]
            digepred_gene_pairs.append(gene_pair_dict)

    ##########################################################################
    # Was there a generated .png file from DIEP
    ##########################################################################

    diep_png_plot_filename = os.path.join('casewide','DIEP','DIEP.png')
    if os.path.exists(os.path.join(case_root_dir, diep_png_plot_filename)):
        website_filelist.append(diep_png_plot_filename)
        LOGGER.info("Integrating DIEP plot: %s", diep_png_plot_filename)
    else:
        LOGGER.warning("DIEP %s missing.  Did you run DIEP?", diep_png_plot_filename)
        diep_png_plot_filename = None

    # Add final things to go on case summary
    case_jinja2.update({'variant_isoform_summaries': variant_isoform_summaries,
                           'genome_headers': genome_headers,
                           'early_or_fail_message': early_or_fail_message,
                           'refreshFlag': args.embed_refresh,
                           'user_message': args.user_message,
                           'refresh_interval_seconds': args.refresh_interval_seconds,
                           'seconds_remaining': args.seconds_remaining,
                           'vus': vustruct,
                           'vustruct_logs_info': vustruct_logs_info,
                           # 'firstGeneTable': html_table_generic,
                           # 'firstGeneReport': html_report_generic,
                           # 'secondGeneTable': html_table_familial,
                           # 'secondGeneReport': html_report_familial,
                           'disease1_variant_short_description': config_pathprox_dict[
                               'disease1_variant_short_description'],
                           'disease2_variant_short_description': config_pathprox_dict[
                               'disease2_variant_short_description'],
                           'digepred_html_filename': digepred_html_filename,
                           'digepred_gene_pairs': digepred_gene_pairs,
                           'diep_png_plot_filename': diep_png_plot_filename
                           })
    if LOGGER.isEnabledFor(logging.DEBUG):
        pp = pprint.PrettyPrinter(indent=1)
        LOGGER.debug("Dictionary case_jinja2 to render:\n%s" % pp.pformat(case_jinja2))

    case_summary_filename = write_case_report_html(case_jinja2)

    # Remove vustruct_logs_info from the structure before we do more with it
    # Because we cannot convert the embedded dataframe to JSON here
    del case_jinja2['vustruct_logs_info']

    last_message = "The case summary report is: " + case_summary_filename
    case_summary_json = os.path.join(case_root_dir,
                                     "%s.json" % args.projectORstructure)

    del case_jinja2['vus']
    with open(case_summary_json, 'w') as fp:
        json.dump(case_jinja2, fp)


    LOGGER.info("Writing website log/ files")

    html_out = vustruct_logs_template.render(vustruct_logs_info)
    vustruct_html_filename = os.path.join("log/", "index.html")
    website_filelist.append(vustruct_html_filename)
    with open(vustruct_html_filename, "w") as f:
        f.write(html_out)

    print("")
    print("")

    if args.verbose:
        LOGGER.info(last_message)
    else:
        print(last_message)


    # - plots for the PPI and ScanNet
    ppi_graph_xaxis_labels = []
    ppi_graph_values = []

    ptm_df_columns = ['Gene', 'Native Residue', 'PTM Residue', 'Type', 'Probability']
    ptm_df_rows = []

    if not genome_headers:
        # Without genomic coordinate input, we populate the summary spreadsheet with each variant row
        # When genomic coordinates are seen replicated, then we summarize (else: block) per row
        for vis in variant_isoform_summaries:
            if len(vis) > 0:
                _min_ScanNetPPI,_max_ScanNetPPI = variant_isoform_helper.min_max_ScanNetPPI([vis])
                _min_MusiteDeepPTM,_max_MusiteDeepPTM = variant_isoform_helper.min_max_MusiteDeepPTM([vis])
                row = [
                    vis['Gene'], 
                    vis['Unp'], 
                    vis['Mutation'], 
                _format_min_max(vis['ddG Cartesian Min'],vis['ddG Cartesian Max']),
                _format_min_max(vis['disease1_pp Min'],vis['disease1_pp Max']),
                _format_min_max(vis['disease2_pp Min'],vis['disease2_pp Max']),
                vis['cosmis_score'],
                _format_min_max(_min_ScanNetPPI,_max_ScanNetPPI,0,True),
                vis['alphamissense_score'],
                _format_min_max(_min_MusiteDeepPTM,_max_MusiteDeepPTM,0,True)]
                case_summary_worksheet.append(row) 

        # End if NOT genome_headers
    else:
        for gh in genome_headers:
            # "Gene","Uniprot ID","Variant", "ddG", "PathProx 3D hotspot vs Clinvar", "PathProx 3D hotspot vs COSMIC",
            # "PPI", "AlphaMissense", "PTM", "Predicted Digenic Interactions","Worth pursuing?","side nodes"]
        
            canonical_unp_if_avail = ""
            variant_if_avail = ""
            vis = gh['variant_isoform_summaries']
            _min_ScanNetPPI = 0.0
            _max_ScanNetPPI = 0.0
            if len(vis) > 0:
                _min_ScanNetPPI,_max_ScanNetPPI = variant_isoform_helper.min_max_ScanNetPPI(vis)
                canonical_unp_if_avail = vis[0]['Unp'].split('-')[0]
                # Later we can use MANE transcripts or others if 
                # the canonical uniprot transcript is unaffected
                # For now, we default to the [0]th transcript in the set
                # for the genomic coordinates
                variant_if_avail = vis[0]['Mutation']
                vis_canonical = vis[0]
                for vi_summary in vis:
                    if PDBMapProtein.isCanonicalByUniparc(vi_summary['Unp']):
                        vis_canonical = vi_summary
                        canonical_unp_if_avail = vi_summary['Unp']
                        variant_if_avail = vi_summary['Mutation']
    
    
                if 'MusiteDeepPTM' in vis_canonical and vis_canonical['MusiteDeepPTM']:
                    for musite_deep_dict in vis_canonical['MusiteDeepPTM']:
                        # Only retain those predictions of higher than 50%
                        if musite_deep_dict['PTM Probability'] > 0.5:
                            ptm_row = [gh['Gene'], 
                                       musite_deep_dict['Native Residue'], 
                                       musite_deep_dict['PTM Residue'], 
                                       musite_deep_dict['PTM Type'], 
                                       musite_deep_dict['PTM Probability']]
                            ptm_df_rows.append(ptm_row)
    
            if variant_if_avail:
                # No need to include the variant AA in the tick.
                _ppi_graph_xaxis_label = '{}\n{}'.format(gh['Gene'],variant_if_avail[0:-1])
            else:
                _ppi_graph_xaxis_label = gh['Gene']
    
            ppi_graph_xaxis_labels.append(_ppi_graph_xaxis_label)

            try:
                # _scannet_ppi = float(gh['ScanNetPPI'])
                _scannet_ppi = float(_max_ScanNetPPI)
            except (ValueError,TypeError) as ve:
                _scannet_ppi = 0.0
            ppi_graph_values.append(_scannet_ppi)
    
            _min_alphamissense_score, _max_alphamissense_score = \
                variant_isoform_helper.min_max_alphamissense(vis)
                
            row = [
                gh['Gene'], 
                canonical_unp_if_avail, 
                variant_if_avail, 
                _format_min_max(gh['ddG Cartesian Min'],gh['ddG Cartesian Max']),
                _format_min_max(gh['disease1_pp Min'],gh['disease1_pp Max']),
                _format_min_max(gh['disease2_pp Min'],gh['disease2_pp Max']),
                gh['cosmis_score'],
                gh['ScanNetPPI'],
                _format_min_max(_min_alphamissense_score, _max_alphamissense_score),
                gh['MusiteDeepPTM']]
            case_summary_worksheet.append(row) 
        # END if genomeheaders:

    # Do final population of the additional outputs and write them
    ddg_populate_worksheet(ddG_cartesian_worksheet, ddG_cartesian_xref)
    ddg_populate_worksheet(ddG_monomer_worksheet, ddG_monomer_xref)
            
    max_row = case_summary_worksheet.max_row
    case_summary_worksheet.cell(row=max_row+2,column=10).value = "legend"

    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
    yellowFill = PatternFill(start_color='FFFFFF00',
                   end_color='FFFFFF00',
                   fill_type='solid')
    blueFill = PatternFill(start_color='FF0000FF',
                   end_color='FF0000FF',
                   fill_type='solid')
    case_summary_worksheet.cell(row=max_row+3,column=10).value = ""
    case_summary_worksheet.cell(row=max_row+3,column=10).fill = redFill
    case_summary_worksheet.cell(row=max_row+3,column=11).value = "Yes"


    case_summary_worksheet.cell(row=max_row+4,column=10).value = ""
    case_summary_worksheet.cell(row=max_row+4,column=10).fill = yellowFill
    case_summary_worksheet.cell(row=max_row+4,column=11).value = "Maybe"

    case_summary_worksheet.cell(row=max_row+5,column=10).value = ""
    case_summary_worksheet.cell(row=max_row+5,column=10).fill = blueFill
    case_summary_worksheet.cell(row=max_row+5,column=11).value = "No"


    _case_workbook_filename = os.path.join(REPORT_FILES_DIRECTORY,args.projectORstructure+"_case.xlsx")
    LOGGER.info("Saving case summary of summary file: %s" , _case_workbook_filename)
    case_workbook.save(_case_workbook_filename)
    website_filelist.append(_case_workbook_filename)

    # Now save the PPI column chart
    plt.bar(ppi_graph_xaxis_labels, ppi_graph_values)
    plt.title('ScanNet Predicted Protein-protein interaction', fontsize=17)
    plt.ylim(0,1.0)
    plt.yticks(fontsize=8)
    plt.xticks(fontsize=8, rotation=45)
    plt.xlabel('Gene and Amino Acid Variant', fontsize=12)
    plt.ylabel('PPI predicted probability', fontsize=12)
    plt.axhline(0.5, color='black')
    plt.tight_layout()
   
    _ppi_plot_filename = os.path.join(REPORT_FILES_DIRECTORY,"PPI_%s.png" % args.projectORstructure)
    LOGGER.info("Saving PPI plot file: %s", _ppi_plot_filename)
    plt.savefig(_ppi_plot_filename, dpi=600)
    website_filelist.append(_ppi_plot_filename)
    # plt.show()

    # Now save the PTM dataframe
    ptm_df = pd.DataFrame(ptm_df_rows,columns=ptm_df_columns)
    _ptm_neighborhood_filename = os.path.join(REPORT_FILES_DIRECTORY,"PTM_%s_neighborhood.csv" % args.projectORstructure)
    LOGGER.info("Saving PTM file: %s", _ptm_neighborhood_filename)
    ptm_df.to_csv(_ptm_neighborhood_filename, index=None)
    website_filelist.append(_ptm_neighborhood_filename)

    write_case_website_tar_zip(case_summary_filename)

sys.exit(0)
